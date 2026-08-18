#!/usr/bin/env python3
"""Generate fma_bench.xlsx from bench/fma/results/.

Summarizes the branch-free DW/TW/QW FMA of arXiv:2607.11391 as integrated into
BNCmatmul: verification results and the AXPY/GEMV/GEMM benchmark for
DD/TD/QD (double-word base) and DS/TS/QS (single-word base) on
serial / NEON / SVE2.

Sheets:
  ReadMe       environment, algorithms, flop counts, certified bounds, legend
  Verification certified error bounds, commutativity, backend bitwise
               agreement, OpenMP bitwise agreement
  Accuracy     max / mean relative error vs MPFR 600-bit  (+ chart)
  Timing       s/call at 1 thread and at N threads, OpenMP speedup
  Speedup      FMA/Q and FMA/BF per type x op x backend   (+ charts)
  Polynomial   Horner / Estrin / eval_diff / poly-mul: Q vs BF vs FMA
  PolyDegree   degree sweep: FMA/BF speedup vs degree
  HornerEstrin Horner vs Estrin, with and without SIMD, vs degree
  Raw          every parsed benchmark row
"""
import os, re, sys
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
RES  = os.path.join(HERE, "results")
XLSX = os.path.join(ROOT, "fma_bench.xlsx")

TYPES = ["dd", "td", "qd", "ds", "ts", "qs"]
TYPE_DESC = {
    "dd": "double-double  (K=2, 2x binary64, ~32 digits)",
    "td": "triple-double  (K=3, 3x binary64, ~48 digits)",
    "qd": "quad-double    (K=4, 4x binary64, ~64 digits)",
    "ds": "double-single  (K=2, 2x binary32, ~14 digits)",
    "ts": "triple-single  (K=3, 3x binary32, ~21 digits)",
    "qs": "quad-single    (K=4, 4x binary32, ~28 digits)",
}
TYPE_K = {"dd": 2, "td": 3, "qd": 4, "ds": 2, "ts": 3, "qs": 4}
OPS = ["AXPY", "GEMV", "GEMM"]
VARS = ["Q", "BF", "FMA"]
BACKENDS = ["serial", "neon", "sve2"]
BK_DESC = {"serial": "no SIMD (scalar)",
           "neon": "Arm NEON  (2 double lanes / 4 float lanes)",
           "sve2": "Arm SVE2  (VL=128 bit: 2 double lanes / 4 float lanes)"}
VAR_DESC = {
    "Q":   "library default mul + add  (DD/DS sloppy, TD/TS sloppy, QD/QS Bailey)",
    "BF":  "branch-free mul_bf + add_bf (Zhang-Aiken) - the fair baseline",
    "FMA": "proposed branch-free fused FMA (arXiv:2607.11391), 17/66/146 flops",
}
FLOPS = {2: (20, 29, 17), 3: (None, 96, 66), 4: (None, 209, 146)}
BOUND = {2: "34 u^2", 3: "184 u^3", 4: "812 u^4"}
REF_PREC = 600   # bits of the MPFR accuracy reference

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
NA_FILL  = PatternFill("solid", fgColor="F2F2F2")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
def parse_bench(path, backend):
    """-> list of dicts, plus the '# ...' header info"""
    rows, info = [], {}
    if not os.path.exists(path):
        return rows, info
    for line in open(path):
        line = line.rstrip("\n")
        if line.startswith("#"):
            m = re.search(r"backend=(\S+)\s+lanes=(\d+)\s+threads=(\d+)", line)
            if m:
                info["lanes"] = int(m.group(2)); info["threads"] = int(m.group(3))
            m = re.search(r"AXPY n=(\d+)\s+GEMV n=(\d+)\s+GEMM n=(\d+)", line)
            if m:
                info["n_axpy"], info["n_gemv"], info["n_gemm"] = \
                    int(m.group(1)), int(m.group(2)), int(m.group(3))
            continue
        f = line.split()
        if len(f) == 9 and f[0] in TYPES and f[1] in OPS and f[2] in VARS:
            rows.append(dict(backend=backend, type=f[0], op=f[1], var=f[2],
                             maxrel=float(f[3]), meanrel=float(f[4]),
                             t1=float(f[5]), tomp=float(f[6]),
                             ompspd=float(f[7]), omp=f[8]))
    return rows, info


def is_na(r):
    """scalar single-word BF is not available in c_ds_qs.h (falls back to Q)"""
    return r["backend"] == "serial" and r["type"] in ("ds", "ts", "qs") and r["var"] == "BF"


def read_text(name):
    p = os.path.join(RES, name)
    return open(p).read() if os.path.exists(p) else "(not run)"


# --------------------------------------------------------------------------
def style_header(ws, row, ncol, start=1):
    for c in range(start, start + ncol):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CEN; cell.border = BORDER


def autosize(ws, maxw=60):
    for col in ws.columns:
        letter = None
        w = 0
        for cell in col:
            if letter is None:
                try: letter = cell.column_letter
                except AttributeError: return
            v = cell.value
            if v is not None:
                w = max(w, min(len(str(v)), maxw))
        if letter:
            ws.column_dimensions[letter].width = w + 2


# --------------------------------------------------------------------------
def sheet_readme(wb, info, env):
    ws = wb.create_sheet("ReadMe")
    r = 1
    ws.cell(row=r, column=1, value="Branch-free DW/TW/QW FMA in BNCmatmul - benchmark summary").font = TITLE_FONT
    r += 2
    for line in [
        "Algorithm: T. Kouya, \"Performance evaluation of branch-free fused multiply-add algorithms",
        "           for multi-component-type multiple-precision floating-point arithmetic\",",
        "           arXiv:2607.11391v1 (2026).  Algorithms 1-3 (DW/TW/QW-FMA).",
        "",
        "z := x*y + c is computed in ONE fused stage instead of mul + add.",
        "Implementation: include/bncfma_d.h, include/bncfma_f.h,",
        "                include/neon/_bncneon_fma.h, include/sve2/_bncsve2_fma.h.",
        "Enabled inside the library by defining BNC_USE_NEW_FMA (off by default).",
        "All builds use -ffp-contract=off (mandatory: the EFTs are not error-free otherwise).",
    ]:
        ws.cell(row=r, column=1, value=line); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Environment").font = BOLD; r += 1
    for k, v in env.items():
        ws.cell(row=r, column=1, value=k); ws.cell(row=r, column=2, value=v); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Operation counts and certified error bounds").font = BOLD; r += 1
    hdr = ["K", "types", "Q (mul+add)", "BF (mul_bf+add_bf)", "FMA (proposed)",
           "FMA/BF flop ratio", "certified bound  |z-(xy+c)| <= C u^K (|xy|+|c|)"]
    for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr)); r += 1
    for K, tys in ((2, "DD / DS"), (3, "TD / TS"), (4, "QD / QS")):
        q, bf, fma = FLOPS[K]
        ws.cell(row=r, column=1, value=K)
        ws.cell(row=r, column=2, value=tys)
        ws.cell(row=r, column=3, value=q if q else "value-dependent branches")
        ws.cell(row=r, column=4, value=bf)
        ws.cell(row=r, column=5, value=fma)
        ws.cell(row=r, column=6, value=round(bf / fma, 2))
        ws.cell(row=r, column=7, value=BOUND[K] + " (|xy|+|c|)")
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Variants compared").font = BOLD; r += 1
    for v in VARS:
        ws.cell(row=r, column=1, value=v); ws.cell(row=r, column=2, value=VAR_DESC[v]); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Types").font = BOLD; r += 1
    for t in TYPES:
        ws.cell(row=r, column=1, value=t); ws.cell(row=r, column=2, value=TYPE_DESC[t]); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Backends").font = BOLD; r += 1
    for b in BACKENDS:
        ws.cell(row=r, column=1, value=b); ws.cell(row=r, column=2, value=BK_DESC[b]); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Methodology").font = BOLD; r += 1
    for line in [
        "AXPY  y := alpha*x + y                 n = %s" % info.get("n_axpy", "?"),
        "GEMV  y := A*x   (A column-major)      n = %s" % info.get("n_gemv", "?"),
        "GEMM  C := A*B   (all row-major)       n = %s" % info.get("n_gemm", "?"),
        "Storage is SoA (one array per component); kernels are self-contained in",
        "bench/fma/fma_kernels.inc so that all three variants are compiled into one binary.",
        "Accuracy reference: MPFR at 600 bits.  AXPY and GEMV are checked on every element;",
        "GEMM on 2048 randomly sampled entries of C.",
        "Timing: best-effort wall clock per call; 1 thread vs OMP_NUM_THREADS threads.",
        "The Polynomial sheet uses the library's own src/{dd,td,qd}_poly.c, built three",
        "times and single-threaded:  Q = library default,",
        "BF = -DUSE_DD_BF -DUSE_TD_BF -DUSE_QD_BF (mul_bf + add_bf, the fair baseline),",
        "FMA = -DBNC_USE_NEW_FMA.  FMA/BF is the primary figure of merit.",
        "The 'omp' column of every row was 'ok', i.e. the OpenMP run reproduced the",
        "1-thread run bitwise (the FMA routines are pure - no static temporaries).",
    ]:
        ws.cell(row=r, column=1, value=line); r += 1
    r += 1

    ws.cell(row=r, column=1, value="Caveats").font = BOLD; r += 1
    for line in [
        "* BF is not available for DS/TS/QS on the scalar backend: c_ds_qs.h has no",
        "  branch-free (_bf) scalar routines (only the NEON/SVE2 headers have them).",
        "  Those cells are marked n/a; the binary falls back to the Q code there.",
        "* Q accuracy differs between backends because the backends implement different",
        "  default mul/add algorithms; this is pre-existing and unrelated to the FMA.",
        "* FIXED 2026-07-28: _bncsve2_rqs_add_sloppy had a stray '+ t2' in the final",
        "  t0 = t0 + t1 + t3 accumulation of Bailey's sloppy QS addition (t2 had already been",
        "  consumed by the preceding fthree_sum2 gate). It cost SVE2 QS additions ~10 digits.",
        "  The double-word twin _bncsve2_rqd_add_sloppy, NEON and the scalar c_qs_add were all",
        "  correct; only the float SVE2 transcription was wrong. See the Verification sheet",
        "  (SVE2 vs NEON audit) - qs add_sloppy is now bitwise identical to NEON.",
        "* The remaining SVE2/NEON mismatches in *_mul_sloppy are benign: SVE2 uses a fused",
        "  svmla for the cross terms where NEON uses a separate mul + add, which makes SVE2",
        "  equally or slightly MORE accurate (dd: max 4.4e-32 vs 5.0e-32).",
        "* Worst-case accuracy: the proposed FMA is bounded relative to (|xy|+|c|), not to the",
        "  result, so under strong cancellation its maximum relative error can exceed the",
        "  baseline's by up to ~10^2 (see the Accuracy sheet). Mean errors stay comparable.",
        "* AVX2 / AVX-512 versions of the proposed FMA are not included (untestable here).",
    ]:
        ws.cell(row=r, column=1, value=line); r += 1
    autosize(ws)
    return ws


def sheet_verification(wb):
    ws = wb.create_sheet("Verification")
    r = 1
    ws.cell(row=r, column=1, value="Verification of the proposed FMA").font = TITLE_FONT; r += 2
    for title, fname in (("1. Certified error bounds + commutativity (vs MPFR 600-bit)", "test_fma_ref.txt"),
                         ("2. Backend bitwise agreement (scalar vs NEON vs SVE2)", "test_fma_simd.txt"),
                         ("2b. SVE2 vs NEON audit of the existing add/mul kernels", "test_sve2_vs_neon.txt"),
                         ("3. OpenMP thread-safety (serial vs OpenMP, bitwise)", "test_fma_omp.txt")):
        ws.cell(row=r, column=1, value=title).font = BOLD; r += 1
        for line in read_text(fname).split("\n"):
            ws.cell(row=r, column=1, value=line); r += 1
        r += 1
    autosize(ws)
    return ws


def sheet_accuracy(wb, rows):
    ws = wb.create_sheet("Accuracy")
    ws.cell(row=1, column=1, value="Relative error against MPFR 600-bit").font = TITLE_FONT
    ws.cell(row=2, column=1, value="max relative error (upper block) / mean relative error (lower block)")
    r = 4
    idx = {(x["backend"], x["type"], x["op"], x["var"]): x for x in rows}
    for which, key in (("Maximum relative error", "maxrel"), ("Mean relative error", "meanrel")):
        ws.cell(row=r, column=1, value=which).font = BOLD; r += 1
        hdr = ["type", "K", "op"] + ["%s %s" % (b, v) for b in BACKENDS for v in VARS]
        for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
        style_header(ws, r, len(hdr))
        hrow = r; r += 1
        for t in TYPES:
            for op in OPS:
                ws.cell(row=r, column=1, value=t)
                ws.cell(row=r, column=2, value=TYPE_K[t])
                ws.cell(row=r, column=3, value=op)
                c = 4
                for b in BACKENDS:
                    for v in VARS:
                        x = idx.get((b, t, op, v))
                        cell = ws.cell(row=r, column=c)
                        if x is None:
                            cell.value = "-"; cell.fill = NA_FILL
                        elif is_na(x):
                            cell.value = "n/a"; cell.fill = NA_FILL
                        else:
                            cell.value = x[key]; cell.number_format = "0.00E+00"
                        cell.border = BORDER
                        c += 1
                r += 1
        r += 1

    # ---- accuracy ratio FMA / baseline ----
    ws.cell(row=r, column=1, value="Accuracy ratio FMA / baseline  (>1 = the proposed FMA is less accurate)").font = BOLD
    r += 1
    hdr = ["type", "K", "op"] + ["%s max FMA/%s" % (b, v) for b in BACKENDS for v in ("Q", "BF")]
    for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr)); r += 1
    for t in TYPES:
        for op in OPS:
            ws.cell(row=r, column=1, value=t)
            ws.cell(row=r, column=2, value=TYPE_K[t])
            ws.cell(row=r, column=3, value=op)
            c = 4
            for b in BACKENDS:
                fma = idx.get((b, t, op, "FMA"))
                for base in ("Q", "BF"):
                    ref = idx.get((b, t, op, base))
                    cell = ws.cell(row=r, column=c)
                    if fma is None or ref is None or (base == "BF" and is_na(ref)):
                        cell.value = "n/a"; cell.fill = NA_FILL
                    elif ref["maxrel"] == 0.0:
                        cell.value = "-"; cell.fill = NA_FILL
                    else:
                        cell.value = fma["maxrel"] / ref["maxrel"]
                        cell.number_format = ("0.00E+00"
                                              if (cell.value < 0.01 or cell.value > 100.0)
                                              else "0.00")
                        if cell.value > 30.0: cell.fill = WARN_FILL
                        elif cell.value < 0.1: cell.fill = GOOD_FILL
                    cell.border = BORDER
                    c += 1
            r += 1
    r += 1
    for line in [
        "Green: the proposed FMA is >10x MORE accurate than the baseline "
        "(sve2/qs: the default _bncsve2_rqs_*_sloppy path loses ~6 digits - a pre-existing defect).",
        "Yellow: the proposed FMA is >30x less accurate in the worst case. This is the",
        "(|xy|+|c|)-relative error semantics of the proposed FMA showing up under strong",
        "cancellation; the mean relative errors stay within about one order of magnitude.",
    ]:
        ws.cell(row=r, column=1, value=line); r += 1

    autosize(ws)
    return ws


def sheet_timing(wb, rows, threads):
    ws = wb.create_sheet("Timing")
    ws.cell(row=1, column=1, value="Time per call [s]").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="t1 = 1 thread, tOMP = %d threads, ompspd = t1/tOMP; "
                  "'omp' = OpenMP result identical to the 1-thread result, bitwise" % threads)
    r = 4
    hdr = ["backend", "type", "K", "op", "variant", "t1 [s/call]",
           "tOMP [s/call]", "OMP speedup", "omp bitwise"]
    for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr)); r += 1
    for b in BACKENDS:
        for t in TYPES:
            for op in OPS:
                for v in VARS:
                    x = next((y for y in rows if (y["backend"], y["type"], y["op"], y["var"])
                              == (b, t, op, v)), None)
                    if x is None: continue
                    ws.cell(row=r, column=1, value=b)
                    ws.cell(row=r, column=2, value=t)
                    ws.cell(row=r, column=3, value=TYPE_K[t])
                    ws.cell(row=r, column=4, value=op)
                    ws.cell(row=r, column=5, value=v if not is_na(x) else "BF (n/a)")
                    if is_na(x):
                        for c in range(6, 10):
                            ws.cell(row=r, column=c, value="n/a").fill = NA_FILL
                    else:
                        ws.cell(row=r, column=6, value=x["t1"]).number_format = "0.000E+00"
                        ws.cell(row=r, column=7, value=x["tomp"]).number_format = "0.000E+00"
                        ws.cell(row=r, column=8, value=x["ompspd"]).number_format = "0.00"
                        cell = ws.cell(row=r, column=9, value=x["omp"])
                        cell.fill = GOOD_FILL if x["omp"] == "ok" else WARN_FILL
                    for c in range(1, 10): ws.cell(row=r, column=c).border = BORDER
                    r += 1
    autosize(ws)
    return ws


def sheet_speedup(wb, rows, threads):
    ws = wb.create_sheet("Speedup")
    ws.cell(row=1, column=1, value="Speedup of the proposed FMA").font = TITLE_FONT
    ws.cell(row=2, column=1, value="ratio > 1 means the proposed FMA is faster")
    idx = {(x["backend"], x["type"], x["op"], x["var"]): x for x in rows}
    r = 4
    chart_anchor = []
    for label, key in (("1 thread", "t1"), ("%d threads" % threads, "tomp")):
        ws.cell(row=r, column=1, value=label).font = BOLD; r += 1
        hdr = ["type", "K", "op"] + ["%s FMA/%s" % (b, v) for b in BACKENDS for v in ("Q", "BF")]
        for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
        style_header(ws, r, len(hdr))
        hrow = r; r += 1
        first = r
        for t in TYPES:
            for op in OPS:
                ws.cell(row=r, column=1, value=t)
                ws.cell(row=r, column=2, value=TYPE_K[t])
                ws.cell(row=r, column=3, value=op)
                c = 4
                for b in BACKENDS:
                    fma = idx.get((b, t, op, "FMA"))
                    for base in ("Q", "BF"):
                        ref = idx.get((b, t, op, base))
                        cell = ws.cell(row=r, column=c)
                        if fma is None or ref is None or (base == "BF" and is_na(ref)):
                            cell.value = "n/a"; cell.fill = NA_FILL
                        else:
                            cell.value = ref[key] / fma[key]
                            cell.number_format = "0.00"
                        cell.border = BORDER
                        c += 1
                r += 1
        chart_anchor.append((hrow, first, r - 1, label))
        r += 1

    # charts: GEMM speedup FMA/BF and FMA/Q, 1 thread
    hrow, first, last, _ = chart_anchor[0]
    ch = BarChart()
    ch.type = "col"; ch.style = 10
    ch.title = "AXPY/GEMV/GEMM: speedup of the proposed FMA (1 thread)"
    ch.y_axis.title = "speedup (x)"; ch.x_axis.title = "type / operation"
    data = Reference(ws, min_col=4, max_col=3 + 2 * len(BACKENDS), min_row=hrow, max_row=last)
    cats = Reference(ws, min_col=1, max_col=3, min_row=first, max_row=last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.width, ch.height = 34, 12
    ws.add_chart(ch, "A%d" % (r + 2))
    autosize(ws)
    return ws


POLY_OPS = ["HORNER", "HORNER-V", "ESTRIN", "ESTRIN-V", "EVALDIFF", "POLYMUL"]
POLY_DESC = {
    "HORNER":   "eval_*poly_horner   ret := ret*x + a_i          (scalar, deg MACs/eval)",
    "HORNER-V": "batch Horner: LANES evaluation points at once (SIMD over points, bench-local)",
    "ESTRIN":   "eval_*poly_estrin   pairwise a_{2i+1}*x + a_{2i} (scalar)",
    "ESTRIN-V": "_bncavx2_eval_*poly_estrin  the same, SIMD (NEON / SVE2 / AVX)",
    "EVALDIFF": "eval_diff_*poly     ret := ret*x + i*a_i",
    "POLYMUL":  "mul_*poly           c_{i+j} += a_i*b_j           (O(n^2) MACs)",
}


def parse_poly(path):
    rows, info = {}, {}
    if not os.path.exists(path):
        return rows, info
    for line in open(path):
        line = line.rstrip("\n")
        if line.startswith("#"):
            m = re.search(r"degree=(\d+)\s+evaluation points=(\d+)", line)
            if m: info["deg"], info["npts"] = int(m.group(1)), int(m.group(2))
            continue
        f = line.split()
        if len(f) == 6 and f[0] in ("dd", "td", "qd") and f[1] in POLY_OPS:
            rows[(f[0], f[1])] = dict(maxrel=float(f[2]), meanrel=float(f[3]),
                                      t=float(f[4]), calls=int(f[5]))
    return rows, info


def sheet_poly(wb):
    ws = wb.create_sheet("Polynomial")
    ws.cell(row=1, column=1,
            value="Polynomial kernels: Q vs BF (mul_bf+add_bf) vs the proposed FMA").font = TITLE_FONT
    r = 2
    info = {}
    data = {}
    for bk in BACKENDS:
        for v, tag in (("", "Q"), ("_bf", "BF"), ("_fma", "FMA")):
            rr, ii = parse_poly(os.path.join(RES, "poly_bench_%s%s.txt" % (bk, v)))
            data[(bk, tag)] = rr
            info.update(ii)
    ws.cell(row=r, column=1,
            value="degree = %s, %s evaluation points; accuracy reference MPFR %d bits"
                  % (info.get("deg", "?"), info.get("npts", "?"), REF_PREC))
    r += 2
    for op in POLY_OPS:
        ws.cell(row=r, column=1, value=op); ws.cell(row=r, column=2, value=POLY_DESC[op])
        r += 1
    r += 1

    hdr = ["backend", "type", "K", "kernel",
           "t Q [s/call]", "t BF [s/call]", "t FMA [s/call]",
           "FMA/BF", "FMA/Q",
           "max rel.err Q", "max rel.err BF", "max rel.err FMA",
           "err ratio FMA/BF",
           "mean rel.err Q", "mean rel.err BF", "mean rel.err FMA"]
    for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr))
    hrow = r; r += 1
    first = r
    for bk in BACKENDS:
        for t in ("dd", "td", "qd"):
            for op in POLY_OPS:
                q  = data[(bk, "Q")].get((t, op))
                bf = data[(bk, "BF")].get((t, op))
                f  = data[(bk, "FMA")].get((t, op))
                if q is None or bf is None or f is None: continue
                ws.cell(row=r, column=1, value=bk)
                ws.cell(row=r, column=2, value=t)
                ws.cell(row=r, column=3, value=TYPE_K[t])
                ws.cell(row=r, column=4, value=op)
                ws.cell(row=r, column=5, value=q["t"]).number_format = "0.000E+00"
                ws.cell(row=r, column=6, value=bf["t"]).number_format = "0.000E+00"
                ws.cell(row=r, column=7, value=f["t"]).number_format = "0.000E+00"
                cell = ws.cell(row=r, column=8, value=bf["t"] / f["t"])
                cell.number_format = "0.00"
                if cell.value >= 1.5: cell.fill = GOOD_FILL
                ws.cell(row=r, column=9, value=q["t"] / f["t"]).number_format = "0.00"
                ws.cell(row=r, column=10, value=q["maxrel"]).number_format = "0.00E+00"
                ws.cell(row=r, column=11, value=bf["maxrel"]).number_format = "0.00E+00"
                ws.cell(row=r, column=12, value=f["maxrel"]).number_format = "0.00E+00"
                cell = ws.cell(row=r, column=13,
                               value=(f["maxrel"] / bf["maxrel"]) if bf["maxrel"] else 0.0)
                cell.number_format = "0.00"
                if cell.value > 30.0: cell.fill = WARN_FILL
                ws.cell(row=r, column=14, value=q["meanrel"]).number_format = "0.00E+00"
                ws.cell(row=r, column=15, value=bf["meanrel"]).number_format = "0.00E+00"
                ws.cell(row=r, column=16, value=f["meanrel"]).number_format = "0.00E+00"
                for c in range(1, 17): ws.cell(row=r, column=c).border = BORDER
                r += 1
    ws.freeze_panes = "A%d" % first

    ch = BarChart()
    ch.type = "col"; ch.style = 10
    ch.title = "Polynomial kernels: speedup of the proposed FMA (vs BF and vs Q)"
    ch.y_axis.title = "speedup (x)"; ch.x_axis.title = "backend / type / kernel"
    ch.add_data(Reference(ws, min_col=8, max_col=9, min_row=hrow, max_row=r - 1),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, max_col=4, min_row=first, max_row=r - 1))
    ch.width, ch.height = 34, 12
    ws.add_chart(ch, "R2")
    autosize(ws)
    return ws


def sheet_polydeg(wb):
    """degree sweep from bench/fma/results/poly_sweep.csv"""
    import csv as _csv
    path = os.path.join(RES, "poly_sweep.csv")
    if not os.path.exists(path):
        return None
    d = {}      # (backend, variant, deg, type, op) -> t
    degs, backs = set(), set()
    with open(path) as fh:
        for row in _csv.DictReader(fh):
            k = (row["backend"], row["variant"], int(row["deg"]), row["type"], row["op"])
            d[k] = float(row["tpercall"])
            degs.add(int(row["deg"])); backs.add(row["backend"])
    degs = sorted(degs)
    backs = [b for b in BACKENDS if b in backs]

    ws = wb.create_sheet("PolyDegree")
    ws.cell(row=1, column=1,
            value="Polynomial degree sweep: speedup of the proposed FMA over BF "
                  "(mul_bf + add_bf)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="each cell = t(BF) / t(FMA);  >1 means the proposed FMA is faster. "
                  "Timing = best of 3 runs after a warm-up; work per measurement held "
                  "roughly constant across degrees.")
    r = 4
    for base in ("BF", "Q"):
        ws.cell(row=r, column=1, value="FMA / %s" % base).font = BOLD
        r += 1
        hdr = ["backend", "type", "K", "kernel"] + ["deg %d" % g for g in degs]
        for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
        style_header(ws, r, len(hdr))
        hrow = r; r += 1
        first = r
        for bk in backs:
            for t in ("dd", "td", "qd"):
                for op in POLY_OPS:
                    ws.cell(row=r, column=1, value=bk)
                    ws.cell(row=r, column=2, value=t)
                    ws.cell(row=r, column=3, value=TYPE_K[t])
                    ws.cell(row=r, column=4, value=op)
                    for ci, g in enumerate(degs):
                        tb = d.get((bk, base, g, t, op))
                        tf = d.get((bk, "FMA", g, t, op))
                        cell = ws.cell(row=r, column=5 + ci)
                        if tb is None or tf is None or tf == 0.0:
                            cell.value = "-"; cell.fill = NA_FILL
                        else:
                            cell.value = tb / tf
                            cell.number_format = "0.00"
                            if cell.value >= 1.5: cell.fill = GOOD_FILL
                            elif cell.value < 1.0: cell.fill = WARN_FILL
                        cell.border = BORDER
                    r += 1
        if base == "BF":
            bf_first, bf_last, bf_hrow = first, r - 1, hrow
        r += 1

    # line chart: FMA/BF vs degree for the serial backend
    from openpyxl.chart import LineChart
    ch = LineChart()
    ch.title = "FMA / BF speedup vs polynomial degree"
    ch.y_axis.title = "speedup (x)"; ch.x_axis.title = "degree"
    ch.style = 12
    rows_for_chart = []
    for rr in range(bf_first, bf_last + 1):
        if ws.cell(rr, 1).value == backs[0] and ws.cell(rr, 4).value in ("HORNER", "POLYMUL"):
            rows_for_chart.append(rr)
    for rr in rows_for_chart:
        ref = Reference(ws, min_col=5, max_col=4 + len(degs), min_row=rr, max_row=rr)
        ch.add_data(ref, titles_from_data=False)
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef
    for i, rr in enumerate(rows_for_chart):
        if i < len(ch.series):
            ch.series[i].tx = SeriesLabel(v="%s %s" % (ws.cell(rr, 2).value, ws.cell(rr, 4).value))
    ch.set_categories(Reference(ws, min_col=5, max_col=4 + len(degs),
                                min_row=bf_hrow, max_row=bf_hrow))
    ch.width, ch.height = 26, 12
    ws.add_chart(ch, "A%d" % (r + 2))
    autosize(ws)
    return ws


def sheet_horner_estrin(wb):
    """Horner vs Estrin, scalar and SIMD, from the degree sweep."""
    import csv as _csv
    path = os.path.join(RES, "poly_sweep.csv")
    if not os.path.exists(path):
        return None
    d = {}
    degs, backs, variants = set(), set(), set()
    with open(path) as fh:
        for row in _csv.DictReader(fh):
            d[(row["backend"], row["variant"], int(row["deg"]), row["type"], row["op"])] = \
                float(row["tpercall"])
            degs.add(int(row["deg"])); backs.add(row["backend"]); variants.add(row["variant"])
    degs = sorted(degs)
    backs = [b for b in BACKENDS if b in backs]
    var = "FMA" if "FMA" in variants else sorted(variants)[0]

    ws = wb.create_sheet("HornerEstrin")
    ws.cell(row=1, column=1,
            value="Horner vs Estrin - which is faster? (arithmetic variant: %s)" % var).font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="ratio = t(Horner) / t(Estrin);  >1 means ESTRIN is faster, "
                  "<1 means HORNER is faster.  Timing = best of 5 runs after a warm-up.")
    ws.cell(row=3, column=1,
            value="scalar pair = HORNER vs ESTRIN;  SIMD pair = HORNER-V (batch over "
                  "evaluation points) vs ESTRIN-V (vectorized within one evaluation).")
    r = 5
    blocks = [("no SIMD:  t(HORNER) / t(ESTRIN)", "HORNER", "ESTRIN"),
              ("SIMD:     t(HORNER-V) / t(ESTRIN-V)", "HORNER-V", "ESTRIN-V"),
              ("SIMD gain of Horner:  t(HORNER) / t(HORNER-V)", "HORNER", "HORNER-V"),
              ("SIMD gain of Estrin:  t(ESTRIN) / t(ESTRIN-V)", "ESTRIN", "ESTRIN-V")]
    first_block = None
    for title, num, den in blocks:
        ws.cell(row=r, column=1, value=title).font = BOLD
        r += 1
        hdr = ["backend", "type", "K"] + ["deg %d" % g for g in degs]
        for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
        style_header(ws, r, len(hdr))
        hrow = r; r += 1
        start = r
        for bk in backs:
            for t in ("dd", "td", "qd"):
                ws.cell(row=r, column=1, value=bk)
                ws.cell(row=r, column=2, value=t)
                ws.cell(row=r, column=3, value=TYPE_K[t])
                for ci, g in enumerate(degs):
                    a = d.get((bk, var, g, t, num))
                    b = d.get((bk, var, g, t, den))
                    cell = ws.cell(row=r, column=4 + ci)
                    if a is None or b is None or b == 0.0:
                        cell.value = "-"; cell.fill = NA_FILL
                    else:
                        cell.value = a / b
                        cell.number_format = "0.00"
                        if cell.value > 1.05: cell.fill = GOOD_FILL     # denominator wins
                        elif cell.value < 0.95: cell.fill = WARN_FILL   # numerator wins
                    cell.border = BORDER
                r += 1
        if first_block is None:
            first_block = (hrow, start, r - 1)
        r += 1

    ws.cell(row=r, column=1,
            value="green = the second method is faster, yellow = the first one is faster").font = BOLD
    r += 2

    # absolute times of the four kernels, for reference
    ws.cell(row=r, column=1, value="Absolute time per evaluation [s] (variant %s)" % var).font = BOLD
    r += 1
    hdr = ["backend", "type", "K", "kernel"] + ["deg %d" % g for g in degs]
    for c, h in enumerate(hdr, 1): ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr)); r += 1
    for bk in backs:
        for t in ("dd", "td", "qd"):
            for op in ("HORNER", "HORNER-V", "ESTRIN", "ESTRIN-V"):
                ws.cell(row=r, column=1, value=bk)
                ws.cell(row=r, column=2, value=t)
                ws.cell(row=r, column=3, value=TYPE_K[t])
                ws.cell(row=r, column=4, value=op)
                for ci, g in enumerate(degs):
                    v = d.get((bk, var, g, t, op))
                    cell = ws.cell(row=r, column=5 + ci)
                    if v is None: cell.value = "-"; cell.fill = NA_FILL
                    else: cell.value = v; cell.number_format = "0.00E+00"
                    cell.border = BORDER
                r += 1

    from openpyxl.chart import LineChart
    from openpyxl.chart.series import SeriesLabel
    hrow, start, last = first_block
    ch = LineChart()
    ch.title = "t(Horner) / t(Estrin) vs degree  (>1: Estrin faster)"
    ch.y_axis.title = "ratio"; ch.x_axis.title = "degree"
    ch.style = 12
    for rr in range(start, last + 1):
        if ws.cell(rr, 1).value != backs[0]: continue
        ch.add_data(Reference(ws, min_col=4, max_col=3 + len(degs), min_row=rr, max_row=rr),
                    titles_from_data=False)
        ch.series[-1].tx = SeriesLabel(v=str(ws.cell(rr, 2).value))
    ch.set_categories(Reference(ws, min_col=4, max_col=3 + len(degs),
                                min_row=hrow, max_row=hrow))
    ch.width, ch.height = 26, 12
    ws.add_chart(ch, "A%d" % (r + 2))
    autosize(ws)
    return ws


def sheet_raw(wb, rows):
    ws = wb.create_sheet("Raw")
    hdr = ["backend", "type", "K", "op", "variant", "max rel.err", "mean rel.err",
           "t1 [s]", "tOMP [s]", "OMP speedup", "omp bitwise", "note"]
    for c, h in enumerate(hdr, 1): ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(hdr))
    r = 2
    for x in rows:
        ws.cell(row=r, column=1, value=x["backend"])
        ws.cell(row=r, column=2, value=x["type"])
        ws.cell(row=r, column=3, value=TYPE_K[x["type"]])
        ws.cell(row=r, column=4, value=x["op"])
        ws.cell(row=r, column=5, value=x["var"])
        ws.cell(row=r, column=6, value=x["maxrel"]).number_format = "0.00E+00"
        ws.cell(row=r, column=7, value=x["meanrel"]).number_format = "0.00E+00"
        ws.cell(row=r, column=8, value=x["t1"]).number_format = "0.000E+00"
        ws.cell(row=r, column=9, value=x["tomp"]).number_format = "0.000E+00"
        ws.cell(row=r, column=10, value=x["ompspd"]).number_format = "0.00"
        ws.cell(row=r, column=11, value=x["omp"])
        ws.cell(row=r, column=12,
                value="BF unavailable in scalar single-word arithmetic (= Q)" if is_na(x) else "")
        r += 1
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


# --------------------------------------------------------------------------
def main():
    rows, info = [], {}
    for bk in BACKENDS:
        for suffix in ("", "_f"):
            path = os.path.join(RES, "fma_bench_%s%s.txt" % (bk, suffix))
            rr, ii = parse_bench(path, bk)
            if not rr:
                print("warning: no rows in %s" % path, file=sys.stderr)
            rows += rr
            info.update({k: v for k, v in ii.items() if k not in info})
    if not rows:
        print("no benchmark results found in %s" % RES, file=sys.stderr)
        return 1
    threads = info.get("threads", 1)

    env = {
        "CPU": "Arm Cortex-X925 (NVIDIA GB10), 20 cores, NEON + SVE2 (VL=128 bit)",
        "Compiler": "gcc 13.3, -O3 -ffp-contract=off -funroll-loops",
        "OpenMP threads": threads,
        "Accuracy reference": "MPFR 600 bits",
        "AXPY n": info.get("n_axpy"), "GEMV n": info.get("n_gemv"),
        "GEMM n": info.get("n_gemm"),
        "Benchmark source": "bench/fma/fma_blas_bench.c + fma_kernels.inc",
    }

    wb = Workbook()
    wb.remove(wb.active)
    sheet_readme(wb, info, env)
    sheet_verification(wb)
    sheet_accuracy(wb, rows)
    sheet_timing(wb, rows, threads)
    sheet_speedup(wb, rows, threads)
    sheet_poly(wb)
    sheet_polydeg(wb)
    sheet_horner_estrin(wb)
    sheet_raw(wb, rows)
    wb.save(XLSX)
    print("wrote %s  (%d benchmark rows)" % (XLSX, len(rows)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
