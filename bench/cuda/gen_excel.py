#!/usr/bin/env python3
"""Generate cuda_bench.xlsx from bench/cuda/out/cuda_results.csv (+ crossover.csv).

Per operation (matmul / matvec / axpy) the sheet shows PIVOT tables
(rows = dimension N, columns = precision) plus line charts, split real vs
complex so each precision is one clean scaling curve (x = dimension N):

    * GPU time [ms] vs N   (log Y)
    * CPU time [ms] vs N   (log Y)   -- OpenMP+SVE2 / serial baseline
    * speedup CPU/GPU vs N

Line style encodes the number family: the double-based precisions
(d/dd/td/qd, cd/cdd/ctd/cqd) are drawn SOLID, the float/single-based ones
(f/ds/ts/qs, cf/cds/cts/cqs) DASHED.  Legends sit to the right (no overlap with
the plot); large table values use exponential notation.
"""
import csv, os, collections
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "out")
CSV  = os.path.join(OUT, "cuda_results.csv")
XCSV = os.path.join(OUT, "crossover.csv")
XLSX = os.path.join(os.path.dirname(os.path.dirname(HERE)), "cuda_bench.xlsx")

OP_ORDER = ["matmul", "matvec", "axpy"]
OP_TITLE = {
    "matmul": "Matrix product  C = A*B",
    "matvec": "Matrix-vector product  y = A*x",
    "axpy":   "Vector scale  c = val*x  (cmul, memory-bound)",
}
REAL = ["d", "f", "dd", "td", "qd", "ds", "ts", "qs"]
CPLX = ["cd", "cf", "cdd", "ctd", "cqd", "cds", "cts", "cqs"]
PREC_ORDER = REAL + CPLX
# double-based precisions are drawn solid, float/single-based dashed
DOUBLE_FAMILY = {"d", "dd", "td", "qd", "cd", "cdd", "ctd", "cqd"}
# 8 distinct series colours (reused for the real and complex charts)
COLORS = ["1F77B4", "FF7F0E", "2CA02C", "D62728",
          "9467BD", "8C564B", "17BECF", "BCBD22"]

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
SECT_FONT = Font(bold=True, size=11, color="1F4E78")
GW = PatternFill("solid", fgColor="C6EFCE")
RW = PatternFill("solid", fgColor="FFC7CE")


def load(path):
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return list(csv.DictReader(f))


def numfmt_for(v):
    """Exponential for large/tiny magnitudes, plain decimal in between."""
    if v is None:
        return "0.00"
    a = abs(v)
    if a == 0:
        return "0"
    if a >= 1000 or a < 1e-3:
        return "0.00E+00"
    if a < 1:
        return "0.0000"
    return "0.00"


def pivot(rows, op, precs, field):
    dims = sorted({int(r["dim"]) for r in rows if r["op"] == op})
    table = collections.defaultdict(dict)
    for r in rows:
        if r["op"] != op or r["prec"] not in precs:
            continue
        d = int(r["dim"])
        if field == "gpu":
            v = float(r["gpu_time"]) * 1e3
        elif field == "cpu":
            v = float(r["cpu_time"]) * 1e3 if r["cpu_kind"] != "none" and float(r["cpu_time"]) > 0 else None
        else:  # speedup
            gpu = float(r["gpu_time"]); cpu = float(r["cpu_time"])
            v = (cpu / gpu) if (gpu > 0 and cpu > 0 and r["cpu_kind"] != "none") else None
        table[r["prec"]][d] = v
    dims = [d for d in dims if any(table[p].get(d) is not None for p in precs)]
    return dims, table


def style_series(chart, precs):
    """Solid line for double-family, dashed for single/float-family; fixed colours."""
    for i, ser in enumerate(chart.series):
        p = precs[i]
        col = COLORS[i % len(COLORS)]
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill=col, w=22000,
                                 prstDash=("solid" if p in DOUBLE_FAMILY else "sysDash"))
        ser.graphicalProperties = gp
        ser.smooth = False


def add_chart(ws, anchor, title, ylabel, left, hr, last, present, logy):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 9.5; chart.width = 19
    chart.x_axis.title = "dimension N"
    chart.y_axis.title = ylabel
    chart.x_axis.delete = False; chart.y_axis.delete = False
    if logy:
        chart.y_axis.scaling.logBase = 10
    data = Reference(ws, min_col=left + 1, max_col=left + len(present), min_row=hr, max_row=last)
    cats = Reference(ws, min_col=left, min_row=hr + 1, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    style_series(chart, present)
    # legend to the right, NOT overlaying the plot; title stays on top
    chart.legend.position = "r"
    chart.legend.overlay = False
    if chart.title and chart.title.overlay is not None:
        chart.title.overlay = False
    ws.add_chart(chart, anchor)


def write_block(ws, top, left, title, ylabel, dims, table, precs, logy=False,
                shade=False, scientific=True):
    present = [p for p in precs if any(table[p].get(d) is not None for d in dims)]
    if not present or not dims:
        return top
    ws.cell(top, left, title).font = SECT_FONT
    hr = top + 1
    h = ws.cell(hr, left, "N \\ prec"); h.fill = HDR_FILL; h.font = HDR_FONT
    for j, p in enumerate(present, 1):
        c = ws.cell(hr, left + j, p)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(horizontal="center")
    for i, d in enumerate(dims, 1):
        rr = hr + i
        ws.cell(rr, left, d).font = Font(bold=True)
        for j, p in enumerate(present, 1):
            v = table[p].get(d)
            cell = ws.cell(rr, left + j, v)
            if shade:
                cell.number_format = '0.00"x"'
                if v is not None:
                    cell.fill = GW if v >= 1.0 else RW
            else:
                cell.number_format = numfmt_for(v) if scientific else "0.0000"
    last = hr + len(dims)
    for j in range(len(present) + 1):
        ws.column_dimensions[get_column_letter(left + j)].width = 10
    anchor = get_column_letter(left + len(present) + 2) + str(top)
    add_chart(ws, anchor, title, ylabel, left, hr, last, present, logy)
    # advance past whichever is taller: the pivot table or the (~20-row) chart,
    # so successive charts in the right-hand column never overlap each other.
    return top + max(len(dims) + 3, 21)


def op_sheet(wb, op, rows):
    data = [r for r in rows if r["op"] == op]
    if not data:
        return
    ws = wb.create_sheet(op)
    ws["A1"] = OP_TITLE.get(op, op); ws["A1"].font = TITLE_FONT
    ws["A2"] = ("GPU = NVIDIA GB10 (sm_121).  CPU = OpenMP+SVE2 / serial.  "
                "Solid = double-based, dashed = float/single-based.  X = dimension N.")
    row = 4
    for grp, gname in [(REAL, "real"), (CPLX, "complex")]:
        d, t = pivot(data, op, grp, "gpu")
        row = write_block(ws, row, 1, f"GPU time [ms] vs N — {gname}", "GPU time [ms]",
                          d, t, grp, logy=True)
    for grp, gname in [(REAL, "real"), (CPLX, "complex")]:
        d, t = pivot(data, op, grp, "cpu")
        row = write_block(ws, row, 1, f"CPU time [ms] vs N — {gname}", "CPU time [ms]",
                          d, t, grp, logy=True)
    for grp, gname in [(REAL, "real"), (CPLX, "complex")]:
        d, t = pivot(data, op, grp, "speedup")
        if d:
            row = write_block(ws, row, 1,
                              f"speedup CPU/GPU vs N — {gname}  (>=1 green = GPU faster)",
                              "speedup (x)", d, t, grp, logy=False, shade=True)


def crossover_sheet(wb, xrows):
    rows = [r for r in xrows if r["op"] == "matmul"]
    by = collections.defaultdict(dict); kind = {}
    for r in rows:
        if r["cpu_kind"] == "none":
            continue
        by[r["prec"]][int(r["dim"])] = float(r["cpu_time"]) / float(r["gpu_time"])
        kind[r["prec"]] = r["cpu_kind"]
    dims = sorted({d for p in by for d in by[p]})
    ws = wb.create_sheet("crossover")
    ws["A1"] = "matmul crossover : speedup = CPU/GPU  (>=1 green = GPU faster)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "CPU = OpenMP+SVE2 (20 threads) where available; serial otherwise (f, cd, cf, cds, cts, cqs)"
    hr = 4
    hdr = ["prec", "CPU"] + [f"N={d}" for d in dims] + ["crossover"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(hr, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    rr = hr + 1
    for p in PREC_ORDER:
        if p not in by:
            continue
        ws.cell(rr, 1, p); ws.cell(rr, 2, kind[p]); cross = None
        for i, d in enumerate(dims):
            cell = ws.cell(rr, 3 + i); v = by[p].get(d)
            if v is None:
                cell.value = "-"
            else:
                cell.value = round(v, 2); cell.number_format = '0.00"x"'
                cell.fill = GW if v >= 1 else RW
                if cross is None and v >= 1:
                    cross = d
        ws.cell(rr, 3 + len(dims), ("N=" + str(cross)) if cross else "> max tested")
        rr += 1
    for c, w in enumerate([6, 8] + [9] * len(dims) + [12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def main():
    rows = load(CSV)
    if rows is None:
        raise SystemExit("missing " + CSV)
    xrows = load(XCSV)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("summary")
    ws["A1"] = "BNCmatmul CUDA benchmark : GPU (NVIDIA GB10, sm_121) vs CPU"
    ws["A1"].font = TITLE_FONT
    for i, t in enumerate([
        "Precisions: real  d f | dd td qd | ds ts qs    complex  cd cf | cdd ctd cqd | cds cts cqs",
        "Operations: matmul (C=A*B), matvec (y=A*x), axpy (c=val*x, memory-bound)",
        "CPU = libbncmatmul-0.24-omp_sve2.a (OpenMP+SVE2, max threads); serial for f/cd/cf/cds/cts/cqs.",
        "Each op sheet: GPU-time, CPU-time and speedup line charts (real & complex, log Y).",
        "Line style: SOLID = double-based (d/dd/td/qd, cd...), DASHED = float/single-based (f/ds/ts/qs, cf...).",
        "Large table values use exponential notation; legends sit to the right of each plot.",
        "Notes: complex axpy has no CPU cmul routine -> GPU-only.  serial complex-single matmul",
        "   (cts/cqs) CPU is capped at N=512 (single-thread O(N^3) blow-up).  crossover sheet = matmul",
        "   speedup matrix.  GPU's naive multi-precision matmul stays slower than SVE2/20-thread CPU to N=1024.",
    ], 3):
        ws[f"A{i}"] = t
    ws.column_dimensions["A"].width = 105
    for op in OP_ORDER:
        op_sheet(wb, op, rows)
    if xrows:
        crossover_sheet(wb, xrows)
    wb.save(XLSX)
    print("wrote", XLSX, "(%d rows%s)" % (len(rows), ", +crossover" if xrows else ""))


if __name__ == "__main__":
    main()
