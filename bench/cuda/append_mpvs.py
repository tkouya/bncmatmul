#!/usr/bin/env python3
"""append_mpvs.py -- append THIS machine's EFT-vs-MPFR-CUDA results to an
existing gpu_mpvs_report.xlsx without touching the sheets already in it.

The workbook produced by gen_mpvs_report.py holds one platform (ReadMe / AXPY /
GEMV / GEMM).  This script reads the CSVs of the current run and adds

    ReadMe (<TAG>) , AXPY (<TAG>) , GEMV (<TAG>) , GEMM (<TAG>)

plus a "Cross-platform" sheet that puts the two platforms side by side (the
drivers, sizes, precisions and timing protocol are identical, so the ratio is a
pure hardware/toolchain comparison).

Usage:
  python3 bench/cuda/append_mpvs.py --xlsx bench/cuda/out/gpu_mpvs_report.xlsx \\
      --tag GB10 --device "NVIDIA GB10 (sm_121)" --base-tag H100
"""
import argparse, csv, os, sys, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))

PRECS = [("double", "double", 53), ("DD", "dd", 106), ("TD", "td", 159), ("QD", "qd", 212),
         ("float", "float", 24), ("DS", "ds", 48), ("TS", "ts", 72), ("QS", "qs", 96)]
# cu_freal<PB> needs PB % 32 == 0, so it is run at the smallest multiple of 32
# that is >= the EFT bit-width (i.e. at least as accurate).
FREAL_BITS = {53: 64, 106: 128, 159: 160, 212: 224, 24: 32, 48: 64, 72: 96, 96: 96}
OPS = [("AXPY", "axpy", [4096, 16384, 65536]),
       ("GEMV", "matvec", [128, 256, 512]),
       ("GEMM", "matmul", [64, 128, 256])]

HDR_FILL = PatternFill("solid", fgColor="305496"); HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE"); WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
TITLE_FONT = Font(bold=True, size=14); BOLD = Font(bold=True); CEN = Alignment(horizontal="center")
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CEN; cell.border = BORDER


def load_csvs(d):
    """same parsing as gen_mpvs_report.load(), but from an arbitrary directory"""
    eft, mpfr = {}, {}
    with open(os.path.join(d, "mpvs_axpy_eft.csv")) as f:
        for r in csv.DictReader(f):
            if int(r["nthreads"]) == 32:
                continue
            eft[("axpy", r["type"], int(r["N"]))] = (float(r["seconds"]), float(r["mflops"]))
    with open(os.path.join(d, "mpvs_dense_eft.csv")) as f:
        for r in csv.DictReader(f):
            eft[(r["op"], r["type"], int(r["N"]))] = (float(r["gpu_sec"]), float(r["gpu_mflops"]))
    with open(os.path.join(d, "mpvs_mpfr.csv")) as f:
        for r in csv.DictReader(f):
            op, N = r["op"], int(r["N"])
            sec, mf = float(r["gpu_sec"]), float(r["gpu_mflops"])
            if r["kind"] == "native":
                eft[(op, r["type"], N)] = (sec, mf)
            else:
                mpfr[(op, int(r["bits"]), N)] = (sec, mf)
    freal = {}
    fp = os.path.join(d, "mpvs_freal.csv")
    if os.path.exists(fp):
        with open(fp) as f:
            for r in csv.DictReader(f):
                freal[(r["op"], int(r["bits"]), int(r["N"]))] = (float(r["gpu_sec"]),
                                                                 float(r["gpu_mflops"]))
    return eft, mpfr, freal


def sheet_readme(wb, tag, device, lines_extra):
    name = "ReadMe (%s)" % tag
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "GPU multiprecision: fixed EFT vs arbitrary MPFR-CUDA (real) -- %s" % tag
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Device : %s" % device,
        "",
        "Same drivers, sizes, precisions and timing protocol as the first platform's",
        "sheets (ReadMe / AXPY / GEMV / GEMM): cuda_axpy_bench (fused EFT AXPY),",
        "gpu_omp_bench (EFT GEMV/GEMM), mpfr_gpu_bench_<PREC> (MPFR-CUDA + native d/f),",
        "driven by bench/cuda/run_mpvs.sh.  Only the GPU and the toolchain differ, so",
        "the Cross-platform sheet is a like-for-like hardware comparison.",
        "",
        "Precision map (EFT type -> bits used for MPFR):",
        "  double=53  DD=106  TD=159  QD=212   float=24  DS=48  TS=72  QS=96",
        "",
        "Operations & sizes:",
        "  AXPY  y=a+alpha*x   N = 4096 / 16384 / 65536      flop = 2 N",
        "  GEMV  y=A*x         N = 128 / 256 / 512           flop = 2 N^2",
        "  GEMM  C=A*B         N = 64 / 128 / 256            flop = 2 N^3",
        "",
        "Timing: GPU kernel only (device-resident inputs, cudaDeviceSynchronize,",
        "        min wall-clock over 5 reps). MFLOPS = flop / gpu_sec / 1e6.",
        "Speedup column = MPFR_sec / EFT_sec.",
        "",
        "Third arm: cu_freal<PB> (mpc_cuda fixed-precision, significand in REGISTERS).",
        "  cu_freal requires PB % 32 == 0, so it cannot hit 53/106/159/212 exactly; it is",
        "  run at the smallest multiple of 32 that is >= the EFT width (recorded in the",
        "  'freal bits' column): 24->32 48->64 53->64 72->96 96->96 106->128 159->160",
        "  212->224, i.e. cu_freal is always at least as accurate as the EFT it faces.",
        "  EFT/freal (x) = freal_sec / EFT_sec ; freal/MPFR (x) = MPFR_sec / freal_sec.",
        "",
        "Source state: BEFORE the branch-free DW/TW/QW FMA (arXiv:2607.11391) is",
        "        switched on -- BNC_USE_NEW_FMA is undefined in this build.",
        "",
    ] + lines_extra
    for i, s in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=s)
    ws.column_dimensions["A"].width = 94
    return ws


def sheet_op(wb, tag, sheet, op, sizes, eft, mpfr, freal):
    name = "%s (%s)" % (sheet, tag)
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "%s: fixed EFT vs MPFR-CUDA (real) on GPU [%s, gpu kernel time, min/5]" % (sheet, tag)
    ws["A1"].font = TITLE_FONT
    hdr = ["precision", "bits", "N", "EFT [s]", "EFT MFLOPS",
           "MPFR [s]", "MPFR MFLOPS", "EFT speedup (x)",
           "freal bits", "freal [s]", "freal MFLOPS",
           "EFT/freal (x)", "freal/MPFR (x)"]
    r0 = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r0, column=c, value=h)
    style_header(ws, r0, 1, len(hdr))
    r = r0 + 1
    for N in sizes:
        for label, etype, bits in PRECS:
            e = eft.get((op, etype, N)); m = mpfr.get((op, bits, N))
            ws.cell(row=r, column=1, value=label).font = BOLD
            ws.cell(row=r, column=1).fill = SUB_FILL
            ws.cell(row=r, column=2, value=bits)
            ws.cell(row=r, column=3, value=N)
            if e:
                ws.cell(row=r, column=4, value=round(e[0], 8))
                ws.cell(row=r, column=5, value=round(e[1], 2))
            if m:
                ws.cell(row=r, column=6, value=round(m[0], 8))
                ws.cell(row=r, column=7, value=round(m[1], 2))
            if e and m and e[0] > 0:
                cell = ws.cell(row=r, column=8, value=round(m[0] / e[0], 2))
                if cell.value >= 2.0: cell.fill = GOOD_FILL
                elif cell.value < 1.0: cell.fill = WARN_FILL
            fb = FREAL_BITS.get(bits)
            fr = freal.get((op, fb, N)) if fb else None
            if fb:
                ws.cell(row=r, column=9, value=fb)
            if fr:
                ws.cell(row=r, column=10, value=round(fr[0], 8))
                ws.cell(row=r, column=11, value=round(fr[1], 2))
                if e and fr[0] > 0:
                    cell = ws.cell(row=r, column=12, value=round(fr[0] / e[0], 2))
                    if cell.value >= 2.0: cell.fill = GOOD_FILL
                    elif cell.value < 1.0: cell.fill = WARN_FILL
                if m and fr[0] > 0:
                    cell = ws.cell(row=r, column=13, value=round(m[0] / fr[0], 2))
                    if cell.value >= 2.0: cell.fill = GOOD_FILL
                    elif cell.value < 1.0: cell.fill = WARN_FILL
            for c in range(1, len(hdr) + 1):
                ws.cell(row=r, column=c).border = BORDER
            r += 1
    for col, w in zip("ABCDEFGHIJKLM",
                      (11, 7, 9, 14, 14, 14, 14, 17, 11, 14, 14, 15, 15)):
        ws.column_dimensions[col].width = w
    return ws


def read_existing(wb, sheet):
    """pull (precision,bits,N) -> (eft_s, eft_mf, mpfr_s, mpfr_mf) from a sheet"""
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    out = {}
    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label:
            continue
        try:
            bits = int(ws.cell(r, 2).value); N = int(ws.cell(r, 3).value)
        except (TypeError, ValueError):
            continue
        vals = [ws.cell(r, c).value for c in (4, 5, 6, 7)]
        out[(label, bits, N)] = tuple(v if isinstance(v, (int, float)) else None for v in vals)
    return out


def sheet_cross(wb, tag, base_tag, eft, mpfr):
    name = "Cross-platform"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "%s vs %s -- same drivers, sizes and protocol" % (base_tag, tag)
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("ratio columns = %s time / %s time;  >1 means %s is faster."
                % (base_tag, tag, tag))
    r = 4
    for sheet, op, sizes in OPS:
        base = read_existing(wb, sheet)
        if not base:
            continue
        ws.cell(row=r, column=1, value=sheet).font = BOLD
        r += 1
        hdr = ["precision", "bits", "N",
               "EFT [s] %s" % base_tag, "EFT [s] %s" % tag, "EFT ratio",
               "MPFR [s] %s" % base_tag, "MPFR [s] %s" % tag, "MPFR ratio",
               "EFT speedup %s" % base_tag, "EFT speedup %s" % tag]
        for c, h in enumerate(hdr, 1):
            ws.cell(row=r, column=c, value=h)
        style_header(ws, r, 1, len(hdr))
        r += 1
        for N in sizes:
            for label, etype, bits in PRECS:
                b = base.get((label, bits, N))
                e = eft.get((op, etype, N)); m = mpfr.get((op, bits, N))
                if b is None and e is None:
                    continue
                ws.cell(row=r, column=1, value=label).font = BOLD
                ws.cell(row=r, column=1).fill = SUB_FILL
                ws.cell(row=r, column=2, value=bits)
                ws.cell(row=r, column=3, value=N)
                be, bm = (b[0], b[2]) if b else (None, None)
                if be: ws.cell(row=r, column=4, value=round(be, 8))
                if e:  ws.cell(row=r, column=5, value=round(e[0], 8))
                if be and e and e[0] > 0:
                    cell = ws.cell(row=r, column=6, value=round(be / e[0], 2))
                    cell.fill = GOOD_FILL if cell.value >= 1.0 else WARN_FILL
                if bm: ws.cell(row=r, column=7, value=round(bm, 8))
                if m:  ws.cell(row=r, column=8, value=round(m[0], 8))
                if bm and m and m[0] > 0:
                    cell = ws.cell(row=r, column=9, value=round(bm / m[0], 2))
                    cell.fill = GOOD_FILL if cell.value >= 1.0 else WARN_FILL
                if b and be and bm and be > 0:
                    ws.cell(row=r, column=10, value=round(bm / be, 2))
                if e and m and e[0] > 0:
                    ws.cell(row=r, column=11, value=round(m[0] / e[0], 2))
                for c in range(1, len(hdr) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1
        r += 1
    for col, w in zip("ABCDEFGHIJK", (11, 7, 9, 15, 15, 11, 15, 15, 11, 16, 16)):
        ws.column_dimensions[col].width = w
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--csvdir", default=HERE)
    ap.add_argument("--tag", required=True, help="short platform tag, e.g. GB10")
    ap.add_argument("--device", required=True)
    ap.add_argument("--base-tag", default="H100", help="tag of the platform already in the workbook")
    ap.add_argument("--note", action="append", default=[])
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print("no such workbook: %s" % args.xlsx, file=sys.stderr); return 1
    eft, mpfr, freal = load_csvs(args.csvdir)
    wb = load_workbook(args.xlsx)
    before = list(wb.sheetnames)

    extra = ["Notes:"] + ["  " + n for n in args.note] if args.note else []
    extra += ["", "Appended %s from %s" % (datetime.date.today().isoformat(), args.csvdir)]
    sheet_readme(wb, args.tag, args.device, extra)
    for sheet, op, sizes in OPS:
        sheet_op(wb, args.tag, sheet, op, sizes, eft, mpfr, freal)
    sheet_cross(wb, args.tag, args.base_tag, eft, mpfr)

    wb.save(args.xlsx)
    print("updated %s" % args.xlsx)
    print("  kept  : %s" % ", ".join(before))
    print("  added : %s" % ", ".join(s for s in wb.sheetnames if s not in before))
    return 0


if __name__ == "__main__":
    sys.exit(main())
