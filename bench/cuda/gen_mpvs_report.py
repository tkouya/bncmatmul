#!/usr/bin/env python3
"""Build gpu_mpvs_report.xlsx : GPU fixed-precision EFT (double, DD, TD, QD,
float, DS, TS, QS) vs arbitrary-precision MPFR-CUDA (real) for AXPY / GEMV /
GEMM, all on the GPU at matched bit-widths.

Inputs (bench/cuda/):
  mpvs_axpy_eft.csv    EFT fused AXPY, all 8 precisions (cuda_axpy_bench, max-thread config)
  mpvs_dense_eft.csv   EFT GEMV/GEMM for dd/td/qd/ds/ts/qs (gpu_omp_bench)
  mpvs_mpfr.csv        native double/float GEMV/GEMM + MPFR-CUDA(real) all ops (mpfr_gpu_bench_*)
Output: bncmatmul-0.24/gpu_mpvs_report.xlsx
"""
import csv, os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
XLSX = os.path.join(ROOT, "gpu_mpvs_report.xlsx")

# display order = user's: double-family then float-family
PRECS = [  # (label, eft_type, bits)
    ("double", "double", 53),
    ("DD", "dd", 106),
    ("TD", "td", 159),
    ("QD", "qd", 212),
    ("float", "float", 24),
    ("DS", "ds", 48),
    ("TS", "ts", 72),
    ("QS", "qs", 96),
]
OPS = [  # (sheet, csv_op, sizes)
    ("AXPY", "axpy",   [4096, 16384, 65536]),
    ("GEMV", "matvec", [128, 256, 512]),
    ("GEMM", "matmul", [64, 128, 256]),
]

HDR_FILL = PatternFill("solid", fgColor="305496"); HDR_FONT = Font(bold=True, color="FFFFFF")
EFT_FILL = PatternFill("solid", fgColor="E2EFDA"); MPF_FILL = PatternFill("solid", fgColor="FCE4D6")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=14); BOLD = Font(bold=True); CEN = Alignment(horizontal="center")
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load():
    eft = {}    # (op,type,N) -> (sec, mflops)
    mpfr = {}   # (op,bits,N) -> (sec, mflops)
    # EFT AXPY (config B = nthreads != 32)
    with open(os.path.join(HERE, "mpvs_axpy_eft.csv")) as f:
        for r in csv.DictReader(f):
            if int(r["nthreads"]) == 32:
                continue
            eft[("axpy", r["type"], int(r["N"]))] = (float(r["seconds"]), float(r["mflops"]))
    # EFT GEMV/GEMM (dd/td/qd/ds/ts/qs)
    with open(os.path.join(HERE, "mpvs_dense_eft.csv")) as f:
        for r in csv.DictReader(f):
            eft[(r["op"], r["type"], int(r["N"]))] = (float(r["gpu_sec"]), float(r["gpu_mflops"]))
    # native double/float (GEMV/GEMM/AXPY) + MPFR-CUDA
    with open(os.path.join(HERE, "mpvs_mpfr.csv")) as f:
        for r in csv.DictReader(f):
            op, N = r["op"], int(r["N"])
            sec, mf = float(r["gpu_sec"]), float(r["gpu_mflops"])
            if r["kind"] == "native":
                eft[(op, r["type"], N)] = (sec, mf)     # native double/float GEMV/GEMM
            else:  # mpfr
                mpfr[(op, int(r["bits"]), N)] = (sec, mf)
    return eft, mpfr


def style_header(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CEN; cell.border = BORDER


def sheet_readme(wb):
    ws = wb.active; ws.title = "ReadMe"
    ws["A1"] = "GPU multiprecision: fixed EFT vs arbitrary MPFR-CUDA (real)"; ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Device : 4x NVIDIA H100 NVL (sm_90), CUDA 13.0, driver 580 (dev 0 used).",
        "",
        "Compared on the GPU, same real operation & size, at matched bit-widths:",
        "  - EFT (fixed precision): gdtq device routines for DD/TD/QD/DS/TS/QS;",
        "    native CUDA kernels for double/float. Fused AXPY, mul_g*matrix_g*vec (GEMV),",
        "    mul_g*matrix_dev (GEMM, naive 1-thread-per-element).",
        "  - MPFR-CUDA (arbitrary precision, REAL): cu_mpfr at PREC bits, one thread per",
        "    output element / row -- the same kernel structure, so the difference is the",
        "    arithmetic. This is the real counterpart of the complex MPC_CUDA benchmark.",
        "",
        "Precision map (EFT type -> bits used for MPFR):",
        "  double=53  DD=106  TD=159  QD=212   float=24  DS=48  TS=72  QS=96",
        "",
        "Operations & sizes:",
        "  AXPY  y=a+alpha*x   N = 4096 / 16384 / 65536      flop = 2 N",
        "  GEMV  y=A*x         N = 128 / 256 / 512           flop = 2 N^2",
        "  GEMM  C=A*B         N = 64 / 128 / 256            flop = 2 N^3",
        "",
        "Timing: GPU kernel only (device-resident inputs, cudaDeviceSynchronize,",
        "        min wall-clock over 5 reps). MFLOPS = flop / gpu_sec / 1e6.",
        "Speedup column = MPFR_sec / EFT_sec  = how many times faster fixed EFT is than",
        "        arbitrary-precision MPFR-CUDA at the same precision & problem.",
        "",
        "Correctness: MPFR results verified vs a host double reference (relerr ~1e-13,",
        "        i.e. the double reference's own rounding; MPFR is more accurate).",
        "",
        "Reading the results:",
        "  - AXPY / GEMV are memory-bound: fixed EFT (few native ops) is 10-1000x faster.",
        "  - GEMM: EFT double/float win hugely; but the gdtq DD/TD/QD/DS/TS/QS GEMM kernel",
        "    is a naive 1-thread-per-element kernel, so vs the (equally naive but highly",
        "    tuned) MPFR-CUDA kernel the multi-component EFT GEMM is only ~1x (sometimes",
        "    slower) -- a known limitation of the EFT GEMM kernel, not of EFT arithmetic.",
    ]
    for i, s in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=s)
    ws.column_dimensions["A"].width = 94
    return ws


def sheet_op(wb, sheet, op, sizes, eft, mpfr):
    ws = wb.create_sheet(sheet)
    ws["A1"] = f"{sheet}: fixed EFT vs MPFR-CUDA (real) on GPU  [gpu kernel time, min/5]"
    ws["A1"].font = TITLE_FONT
    hdr = ["precision", "bits", "N", "EFT [s]", "EFT MFLOPS",
           "MPFR [s]", "MPFR MFLOPS", "EFT speedup (x)"]
    r0 = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r0, column=c, value=h)
    style_header(ws, r0, 1, len(hdr))
    r = r0 + 1
    maxN = sizes[-1]
    helper = {}   # label -> (eft_mf, mpfr_mf, speedup) at maxN
    for N in sizes:
        for label, etype, bits in PRECS:
            e = eft.get((op, etype, N)); m = mpfr.get((op, bits, N))
            ws.cell(row=r, column=1, value=label).font = BOLD
            ws.cell(row=r, column=1).fill = SUB_FILL
            ws.cell(row=r, column=2, value=bits)
            ws.cell(row=r, column=3, value=N)
            if e:
                ws.cell(row=r, column=4, value=round(e[0], 9)).number_format = "0.00E+00"
                ws.cell(row=r, column=4).fill = EFT_FILL
                ws.cell(row=r, column=5, value=round(e[1], 2)).number_format = "0.00"
                ws.cell(row=r, column=5).fill = EFT_FILL
            if m:
                ws.cell(row=r, column=6, value=round(m[0], 9)).number_format = "0.00E+00"
                ws.cell(row=r, column=6).fill = MPF_FILL
                ws.cell(row=r, column=7, value=round(m[1], 2)).number_format = "0.00"
                ws.cell(row=r, column=7).fill = MPF_FILL
            if e and m and e[0] > 0:
                sp = m[0] / e[0]
                ws.cell(row=r, column=8, value=round(sp, 2)).number_format = "0.00"
                if N == maxN:
                    helper[label] = (e[1], m[1], sp)
            for c in range(1, len(hdr) + 1):
                ws.cell(row=r, column=c).border = BORDER
            r += 1
        r += 1  # blank line between N blocks
    widths = [10, 6, 8, 12, 12, 12, 12, 15]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- chart helper block (largest N) ----
    hbase = r + 1
    ws.cell(row=hbase, column=1, value=f"chart data @ N={maxN}").font = BOLD
    hh = hbase + 1
    for c, h in enumerate(["precision", "EFT MFLOPS", "MPFR MFLOPS", "EFT speedup (x)"], 1):
        ws.cell(row=hh, column=c, value=h)
    style_header(ws, hh, 1, 4)
    for i, (label, etype, bits) in enumerate(PRECS):
        rr = hh + 1 + i
        ws.cell(row=rr, column=1, value=label)
        if label in helper:
            ws.cell(row=rr, column=2, value=round(helper[label][0], 2))
            ws.cell(row=rr, column=3, value=round(helper[label][1], 2))
            ws.cell(row=rr, column=4, value=round(helper[label][2], 3))
    hlast = hh + len(PRECS)

    # chart 1: throughput EFT vs MPFR (log)
    ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
    ch.title = f"{sheet} throughput @ N={maxN}: EFT vs MPFR-CUDA"
    ch.y_axis.title = "MFLOPS (log)"; ch.x_axis.title = "precision"
    ch.y_axis.scaling.logBase = 10
    ch.x_axis.delete = False; ch.y_axis.delete = False
    data = Reference(ws, min_col=2, max_col=3, min_row=hh, max_row=hlast)
    cats = Reference(ws, min_col=1, min_row=hh + 1, max_row=hlast)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    ch.legend.position = "b"; ch.height = 8.5; ch.width = 17
    ws.add_chart(ch, "J3")

    # chart 2: EFT speedup over MPFR (log)
    ch2 = BarChart(); ch2.type = "col"; ch2.grouping = "clustered"
    ch2.title = f"{sheet}: EFT speedup over MPFR-CUDA @ N={maxN}  (x)"
    ch2.y_axis.title = "EFT faster by (x, log)"; ch2.x_axis.title = "precision"
    ch2.y_axis.scaling.logBase = 10
    ch2.x_axis.delete = False; ch2.y_axis.delete = False
    d2 = Reference(ws, min_col=4, max_col=4, min_row=hh, max_row=hlast)
    ch2.add_data(d2, titles_from_data=True); ch2.set_categories(cats)
    ch2.legend.position = "b"; ch2.height = 8.5; ch2.width = 17
    ws.add_chart(ch2, "J21")
    return ws


def main():
    eft, mpfr = load()
    wb = Workbook()
    sheet_readme(wb)
    for sheet, op, sizes in OPS:
        sheet_op(wb, sheet, op, sizes, eft, mpfr)
    wb.save(XLSX)
    print("wrote", XLSX)
    # quick console summary at largest N
    for sheet, op, sizes in OPS:
        N = sizes[-1]; print(f"\n{sheet} @ N={N}:  precision  EFT_MFLOPS  MPFR_MFLOPS  EFTx")
        for label, etype, bits in PRECS:
            e = eft.get((op, etype, N)); m = mpfr.get((op, bits, N))
            if e and m:
                print(f"  {label:7s} {e[1]:12.1f} {m[1]:12.1f} {m[0]/e[0]:8.1f}")


if __name__ == "__main__":
    main()
