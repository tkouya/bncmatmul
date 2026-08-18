#!/usr/bin/env python3
"""Generate matmul_bench.xlsx from bench/mm/out/mm_results.csv.

Sheets:
  ReadMe    - environment, methodology, algorithm/precision legend, notes
  Timing    - long-format table: precision x algorithm x backend, t@{256,512,1024}s
              + effective GFLOP/s (2 n^3 / t) and NEON/SVE2 speedup vs serial
  Accuracy  - per-element max relative error (dim 256, MPFR 512-bit reference)
              with a log-scale grouped bar chart per precision
  Charts    - grouped bar charts of matmul time (dim 1024) per precision
"""
import csv, os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "out")
CSV  = os.path.join(OUT, "mm_results.csv")
XLSX = os.path.join(os.path.dirname(os.path.dirname(HERE)), "matmul_bench.xlsx")

PRECS = ["double", "dd", "td", "qd", "float", "ds", "ts", "qs"]
PREC_DESC = {
    "double": "IEEE binary64 (~16 digits)",
    "dd": "double-double, 2x binary64 (~32 digits)",
    "td": "triple-double, 3x binary64 (~48 digits)",
    "qd": "quad-double, 4x binary64 (~64 digits)",
    "float": "IEEE binary32 (~7 digits)",
    "ds": "double-single, 2x binary32 (~14 digits)",
    "ts": "triple-single, 3x binary32 (~21 digits)",
    "qs": "quad-single, 4x binary32 (~28 digits)",
}
ALGOS = ["triple", "block", "strassen", "winograd"]
ALGO_DESC = {
    "triple": "Triple loop (naive O(n^3))",
    "block": "Blocked / tiled O(n^3)",
    "strassen": "Strassen recursive (O(n^2.807))",
    "winograd": "Strassen-Winograd variant (7 mul, 15 add)",
}
BACKENDS = ["serial", "neon", "sve2"]
BK_DESC = {"serial": "No SIMD", "neon": "Arm NEON (Cortex-A76)",
           "sve2": "Arm SVE2 (Neoverse-V2)"}
DIMS = [256, 512, 1024]

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
NA_FILL  = PatternFill("solid", fgColor="F2F2F2")
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load():
    t = {}   # (prec,backend,algo,dim) -> sec
    a = {}   # (prec,algo) -> max_rel_err  (dim fixed)
    with open(CSV) as f:
        for r in csv.DictReader(f):
            k = r["kind"]
            if k == "time":
                t[(r["prec"], r["backend"], r["algo"], int(r["dim"]))] = float(r["value"])
            elif k == "acc":
                a[(r["prec"], r["algo"])] = float(r["value"])
    return t, a


def style_header(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CEN; cell.border = BORDER


def gflops(n, t):
    return (2.0 * n ** 3) / t / 1e9 if t and t > 0 else None


def sheet_readme(wb):
    ws = wb.active; ws.title = "ReadMe"
    ws["A1"] = "Matrix Multiplication Benchmark - BNCmatmul 0.24"; ws["A1"].font = TITLE_FONT
    rows = [
        "",
        "Scope: 4 algorithms x 3 SIMD backends x 8 precisions, dense square C = A*B.",
        "",
        "Algorithms:",
    ] + [f"   - {k:9s}: {v}" for k, v in ALGO_DESC.items()] + [
        "",
        "SIMD backends (chosen at link time):",
    ] + [f"   - {k:7s}: {v}" for k, v in BK_DESC.items()] + [
        "",
        "Precisions:",
    ] + [f"   - {k:7s}: {v}" for k, v in PREC_DESC.items()] + [
        "",
        "Timing: dims 256 / 512 / 1024 (powers of two, required by Strassen/Winograd).",
        "        Each op auto-repeated until >= 0.20 s; time reported per call.",
        "        Effective GFLOP/s uses the naive count 2 n^3 (so throughput is",
        "        directly comparable across algorithms).",
        "",
        "Accuracy: per-element maximum relative error of C = A*B at n = 256, versus a",
        "          512-bit MPFR reference product built from each precision's actually-",
        "          stored (rich, irrational) inputs. Backend-independent -> measured on",
        "          serial. This isolates each algorithm's rounding-error growth.",
        "",
        "Library algorithms (double, dd, td, qd, ds, ts, qs) come from BNCmatmul",
        "(mul_Xmatrix / _block / _strassen / _winograd_even). 'float' has only a",
        "triple-loop in the library, so block/Strassen/Winograd for float are",
        "implemented in the benchmark harness (auto-vectorized by the compiler).",
        "",
        "NOTE: ds / ts / qs on NEON initially failed to link due to two library bugs,",
        "now FIXED so all 8 precisions have all 3 backends:",
        "  1) {ds,ts,qs}linear.c called _bncneon_r*_sum128 / _abssum128 (missing the",
        "     trailing 'f'); the NEON headers define the ..._sum128f / _abssum128f",
        "     variants (identical signature). -> renamed the 9 call sites.",
        "  2) DS only: include/neon/_bncneon_f.h is a stale, incomplete duplicate of",
        "     _bncneon_ds.h that shares its guard (__BNCNEON_DS_H); included first in",
        "     bncneon.h it shadowed the complete ds header, dropping the reductions",
        "     sum128f/abssum128f/norm128f/abs. -> disabled the stale include.",
        "The 3 affected NEON objects were recompiled and swapped into",
        "libbncmatmul-0.24_neon.a (originals kept as *.premmfix). NEON accuracy matches",
        "serial (reduction-order differences only), confirming the fix is correct.",
    ]
    for i, line in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 92
    return ws


def sheet_timing(wb, t):
    ws = wb.create_sheet("Timing")
    ws["A1"] = "Timing  [seconds per C=A*B call]  and effective GFLOP/s (2 n^3 / t)"
    ws["A1"].font = TITLE_FONT
    hdr = ["Precision", "Algorithm", "Backend"]
    for d in DIMS:
        hdr.append(f"t@{d} [s]")
    hdr += ["GFLOP/s@1024", "Speedup@1024 (vs serial)"]
    r0 = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r0, column=c, value=h)
    style_header(ws, r0, 1, len(hdr))
    r = r0 + 1
    for p in PRECS:
        pstart = r
        for algo in ALGOS:
            for bk in BACKENDS:
                ws.cell(row=r, column=1, value=p)
                ws.cell(row=r, column=2, value=algo)
                ws.cell(row=r, column=3, value=bk)
                for c, d in enumerate(DIMS, start=4):
                    v = t.get((p, bk, algo, d))
                    cell = ws.cell(row=r, column=c)
                    if v is None:
                        cell.value = "N/A"; cell.fill = NA_FILL; cell.alignment = CEN
                    else:
                        cell.value = round(v, 9); cell.number_format = "0.00E+00"
                # GFLOPS @1024
                v1024 = t.get((p, bk, algo, 1024))
                g = gflops(1024, v1024) if v1024 else None
                gc = ws.cell(row=r, column=4 + len(DIMS))
                if g is None:
                    gc.value = "N/A"; gc.fill = NA_FILL; gc.alignment = CEN
                else:
                    gc.value = round(g, 2); gc.number_format = "0.00"
                # this row's own speedup @1024 vs the serial baseline
                ser = t.get((p, "serial", algo, 1024))
                own = t.get((p, bk, algo, 1024))
                sc = ws.cell(row=r, column=5 + len(DIMS))
                if ser and own:
                    sc.value = round(ser / own, 2); sc.number_format = "0.00"
                else:
                    sc.value = "N/A"; sc.fill = NA_FILL; sc.alignment = CEN
                for c in range(1, len(hdr) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1
        # shade precision block header cell
        for rr in range(pstart, r):
            ws.cell(row=rr, column=1).fill = SUB_FILL
            ws.cell(row=rr, column=1).font = BOLD
    widths = [10, 10, 9, 12, 12, 12, 13, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def sheet_accuracy(wb, a):
    ws = wb.create_sheet("Accuracy")
    ws["A1"] = "Max relative error per element  (n=256, vs 512-bit MPFR reference)"
    ws["A1"].font = TITLE_FONT
    r0 = 3
    ws.cell(row=r0, column=1, value="Precision")
    for c, algo in enumerate(ALGOS, start=2):
        ws.cell(row=r0, column=c, value=algo)
    style_header(ws, r0, 1, 1 + len(ALGOS))
    r = r0 + 1
    for p in PRECS:
        ws.cell(row=r, column=1, value=p).font = BOLD
        ws.cell(row=r, column=1).fill = SUB_FILL
        for c, algo in enumerate(ALGOS, start=2):
            v = a.get((p, algo))
            cell = ws.cell(row=r, column=c)
            if v is None:
                cell.value = "N/A"; cell.fill = NA_FILL
            else:
                cell.value = v; cell.number_format = "0.00E+00"
            cell.border = BORDER
        ws.cell(row=r, column=1).border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 11
    for c in range(2, 2 + len(ALGOS)):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # log-scale grouped bar chart (categories = precisions, series = algorithms)
    chart = BarChart(); chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Max relative error by precision & algorithm (log scale)"
    chart.y_axis.title = "max rel. error"; chart.x_axis.title = "precision"
    chart.y_axis.scaling.logBase = 10
    data = Reference(ws, min_col=2, max_col=1 + len(ALGOS), min_row=r0, max_row=r - 1)
    cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9; chart.width = 22
    ws.add_chart(chart, f"A{r + 2}")
    return ws


def sheet_charts(wb, t):
    ws = wb.create_sheet("Charts")
    ws["A1"] = "matmul time [s] at n=1024 (helper data below, charts to the right)"
    ws["A1"].font = TITLE_FONT
    # helper table: one block per precision (rows=algos, cols=backends)
    anchor_row = 3
    chart_positions = []
    for pi, p in enumerate(PRECS):
        base = anchor_row + pi * (len(ALGOS) + 3)
        ws.cell(row=base, column=1, value=p).font = BOLD
        ws.cell(row=base + 1, column=1, value="algo")
        for c, bk in enumerate(BACKENDS, start=2):
            ws.cell(row=base + 1, column=c, value=bk)
        style_header(ws, base + 1, 1, 1 + len(BACKENDS))
        for ri, algo in enumerate(ALGOS):
            rr = base + 2 + ri
            ws.cell(row=rr, column=1, value=algo)
            for c, bk in enumerate(BACKENDS, start=2):
                v = t.get((p, bk, algo, 1024))
                cell = ws.cell(row=rr, column=c)
                if v is None:
                    cell.value = None
                else:
                    cell.value = round(v, 6); cell.number_format = "0.000E+00"
        chart = BarChart(); chart.type = "col"; chart.grouping = "clustered"
        chart.title = f"{p}: matmul time [s], n=1024"
        chart.y_axis.title = "seconds"; chart.x_axis.title = "algorithm"
        data = Reference(ws, min_col=2, max_col=1 + len(BACKENDS),
                         min_row=base + 1, max_row=base + 1 + len(ALGOS))
        cats = Reference(ws, min_col=1, min_row=base + 2, max_row=base + 1 + len(ALGOS))
        chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
        chart.height = 7; chart.width = 12
        ws.add_chart(chart, f"F{base}")
    ws.column_dimensions["A"].width = 11
    return ws


def main():
    t, a = load()
    wb = Workbook()
    sheet_readme(wb)
    sheet_timing(wb, t)
    sheet_accuracy(wb, a)
    sheet_charts(wb, t)
    wb.save(XLSX)
    print("wrote", XLSX)
    print(f"  timing rows: {len(t)}   accuracy rows: {len(a)}")


if __name__ == "__main__":
    main()
