#!/usr/bin/env python3
"""Regenerate matmul_bench.xlsx with BOTH machines:

  Arm  (Neoverse-V2 / Cortex-A76)   backends: serial / neon   / sve2
  x86  (Intel Xeon Gold 6526Y)      backends: serial / avx2   / avx512

The Arm results are the ones already in the workbook (from bench/mm/out/
mm_results_arm.csv); the x86 results (mm_results_x86.csv) are appended.  Every
data sheet carries a "Machine" column so the two hosts sit side by side.

Sheets:
  ReadMe    - environment, methodology, algorithm/precision/backend legend
  Timing    - precision x algorithm x backend, per machine; t@{256,512,1024}s
              + effective GFLOP/s (2 n^3 / t) and speedup vs that host's serial
  Accuracy  - per-element max relative error (dim 256, MPFR 512-bit reference)
  Charts    - per precision, one grouped bar chart comparing all 6 backends
              (both machines) at n=1024
"""
import csv, os
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "out")
XLSX = os.path.join(os.path.dirname(os.path.dirname(HERE)), "matmul_bench.xlsx")

PRECS = ["double", "dd", "td", "qd", "float", "ds", "ts", "qs"]
PREC_DESC = {
    "double": "IEEE binary64 (~16 digits)",
    "dd": "double-double, 2x binary64 (~32 digits)",
    "td": "triple-double, 3x binary64 (~48 digits)",
    "qd": "quad-double, 4x binary64 (~64 digits)",
    "float": "IEEE binary32 (~7 digits)",
    "ds": "double-single, 2x binary32 (~14 digits)",
    "ts": "triple-single, 3x binary32 (~21 digits)",
    "qs": "quad-single, 4x binary32 (~28 digits)",
}
ALGOS = ["triple", "block", "strassen", "winograd"]
ALGO_DESC = {
    "triple": "Triple loop (naive O(n^3))",
    "block": "Blocked / tiled O(n^3)",
    "strassen": "Strassen recursive (O(n^2.807))",
    "winograd": "Strassen-Winograd variant (7 mul, 15 add)",
}
DIMS = [256, 512, 1024]

# machine registry (order = sheet order); csv relative to OUT
MACHINES = [
    {"key": "Arm", "label": "Arm (Neoverse-V2 / Cortex-A76)",
     "csv": "mm_results_arm.csv", "backends": ["serial", "neon", "sve2"],
     "bk_desc": {"serial": "No SIMD",
                 "neon": "Arm NEON (Cortex-A76)",
                 "sve2": "Arm SVE2 (Neoverse-V2)"}},
    {"key": "x86", "label": "x86-64 (Intel Xeon Gold 6526Y)",
     "csv": "mm_results_x86.csv", "backends": ["serial", "avx2", "avx512"],
     "bk_desc": {"serial": "No SIMD (scalar x86-64)",
                 "avx2": "AVX2 + FMA, 256-bit  (-mavx2 -mfma)",
                 "avx512": "AVX-512F + FMA, 512-bit  (-mavx512f -mfma)"}},
]

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
ARM_FILL = PatternFill("solid", fgColor="E2EFDA")   # green tint
X86_FILL = PatternFill("solid", fgColor="FCE4D6")   # orange tint
NA_FILL  = PatternFill("solid", fgColor="F2F2F2")
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MACH_FILL = {"Arm": ARM_FILL, "x86": X86_FILL}


def load(csv_path):
    t, a = {}, {}
    if not os.path.exists(csv_path):
        return t, a
    with open(csv_path) as f:
        for r in csv.DictReader(f):
            if r["kind"] == "time":
                t[(r["prec"], r["backend"], r["algo"], int(r["dim"]))] = float(r["value"])
            elif r["kind"] == "acc":
                a[(r["prec"], r["algo"])] = float(r["value"])
    return t, a


# load every machine's data up front
for m in MACHINES:
    m["t"], m["a"] = load(os.path.join(OUT, m["csv"]))


def style_header(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CEN; cell.border = BORDER


def gflops(n, t):
    return (2.0 * n ** 3) / t / 1e9 if t and t > 0 else None


def sheet_readme(wb):
    ws = wb.active; ws.title = "ReadMe"
    ws["A1"] = "Matrix Multiplication Benchmark - BNCmatmul 0.24"; ws["A1"].font = TITLE_FONT
    rows = [
        "",
        "Scope: 4 algorithms x 8 precisions x 2 machines (3 SIMD backends each),",
        "       dense square C = A*B.",
        "",
        "Machines / SIMD backends (backend chosen at link time):",
    ]
    for m in MACHINES:
        rows.append(f"   {m['label']}:")
        rows += [f"      - {k:7s}: {m['bk_desc'][k]}" for k in m["backends"]]
    rows += [
        "",
        "Algorithms:",
    ] + [f"   - {k:9s}: {v}" for k, v in ALGO_DESC.items()] + [
        "",
        "Precisions:",
    ] + [f"   - {k:7s}: {v}" for k, v in PREC_DESC.items()] + [
        "",
        "Timing: dims 256 / 512 / 1024 (powers of two, required by Strassen/Winograd).",
        "        Each op auto-repeated until >= 0.20 s; time reported per call.",
        "        Effective GFLOP/s uses the naive count 2 n^3 (so throughput is",
        "        directly comparable across algorithms).  Speedup is vs the SAME",
        "        machine's serial backend.",
        "",
        "Accuracy: per-element maximum relative error of C = A*B at n = 256, versus a",
        "          512-bit MPFR reference product built from each precision's actually-",
        "          stored (rich, irrational) inputs. Backend-independent -> measured on",
        "          serial. This isolates each algorithm's rounding-error growth.",
        "",
        "Library algorithms (double, dd, td, qd, ds, ts, qs) come from BNCmatmul",
        "(mul_Xmatrix / _block / _strassen / _winograd_even). 'float' has only a",
        "triple-loop in the library, so block/Strassen/Winograd for float are",
        "implemented in the benchmark harness (auto-vectorized by the compiler).",
        "",
        "Build: gcc -O3 -ffp-contract=off; x86 SIMD via -mavx2 -mfma / -mavx512f -mfma",
        "       (__AVX2__ / __AVX512F__ select the SIMD path in the headers).",
        "x86 host: Intel Xeon Gold 6526Y, Ubuntu 24.04, gcc 13, kernel 6.8;",
        "          x86 serial library rebuilt without -mavx (libbncmatmul-0.24_serialx86.a).",
        "",
        "NOTE (Arm): ds / ts / qs on NEON initially failed to link due to two library",
        "bugs, now FIXED so all 8 precisions have all 3 backends:",
        "  1) {ds,ts,qs}linear.c called _bncneon_r*_sum128 / _abssum128 (missing the",
        "     trailing 'f'); the NEON headers define ..._sum128f / _abssum128f. -> renamed.",
        "  2) DS only: include/neon/_bncneon_f.h is a stale duplicate of _bncneon_ds.h",
        "     sharing its guard; it shadowed the complete ds header. -> disabled include.",
    ]
    for i, line in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 92
    return ws


def sheet_timing(wb):
    ws = wb.create_sheet("Timing")
    ws["A1"] = "Timing  [seconds per C=A*B call]  and effective GFLOP/s (2 n^3 / t)"
    ws["A1"].font = TITLE_FONT
    hdr = ["Machine", "Precision", "Algorithm", "Backend"]
    hdr += [f"t@{d} [s]" for d in DIMS]
    hdr += ["GFLOP/s@1024", "Speedup@1024 (vs serial)"]
    r0 = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r0, column=c, value=h)
    style_header(ws, r0, 1, len(hdr))
    r = r0 + 1
    for m in MACHINES:
        t = m["t"]
        for p in PRECS:
            for algo in ALGOS:
                for bk in m["backends"]:
                    ws.cell(row=r, column=1, value=m["key"]).fill = MACH_FILL[m["key"]]
                    ws.cell(row=r, column=1).font = BOLD
                    ws.cell(row=r, column=2, value=p)
                    ws.cell(row=r, column=3, value=algo)
                    ws.cell(row=r, column=4, value=bk)
                    for c, d in enumerate(DIMS, start=5):
                        v = t.get((p, bk, algo, d))
                        cell = ws.cell(row=r, column=c)
                        if v is None:
                            cell.value = "N/A"; cell.fill = NA_FILL; cell.alignment = CEN
                        else:
                            cell.value = round(v, 9); cell.number_format = "0.00E+00"
                    v1024 = t.get((p, bk, algo, 1024))
                    g = gflops(1024, v1024) if v1024 else None
                    gc = ws.cell(row=r, column=5 + len(DIMS))
                    if g is None:
                        gc.value = "N/A"; gc.fill = NA_FILL; gc.alignment = CEN
                    else:
                        gc.value = round(g, 2); gc.number_format = "0.00"
                    ser = t.get((p, "serial", algo, 1024))
                    own = t.get((p, bk, algo, 1024))
                    sc = ws.cell(row=r, column=6 + len(DIMS))
                    if ser and own:
                        sc.value = round(ser / own, 2); sc.number_format = "0.00"
                    else:
                        sc.value = "N/A"; sc.fill = NA_FILL; sc.alignment = CEN
                    for c in range(1, len(hdr) + 1):
                        ws.cell(row=r, column=c).border = BORDER
                    r += 1
    widths = [8, 10, 10, 9, 12, 12, 12, 13, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def sheet_accuracy(wb):
    ws = wb.create_sheet("Accuracy")
    ws["A1"] = "Max relative error per element  (n=256, vs 512-bit MPFR reference)"
    ws["A1"].font = TITLE_FONT
    r0 = 3
    ws.cell(row=r0, column=1, value="Machine")
    ws.cell(row=r0, column=2, value="Precision")
    for c, algo in enumerate(ALGOS, start=3):
        ws.cell(row=r0, column=c, value=algo)
    style_header(ws, r0, 1, 2 + len(ALGOS))
    r = r0 + 1
    for m in MACHINES:
        a = m["a"]
        for p in PRECS:
            ws.cell(row=r, column=1, value=m["key"]).fill = MACH_FILL[m["key"]]
            ws.cell(row=r, column=1).font = BOLD
            ws.cell(row=r, column=2, value=p).font = BOLD
            ws.cell(row=r, column=2).fill = SUB_FILL
            for c, algo in enumerate(ALGOS, start=3):
                v = a.get((p, algo))
                cell = ws.cell(row=r, column=c)
                if v is None:
                    cell.value = "N/A"; cell.fill = NA_FILL
                else:
                    cell.value = v; cell.number_format = "0.00E+00"
                cell.border = BORDER
            ws.cell(row=r, column=1).border = BORDER
            ws.cell(row=r, column=2).border = BORDER
            r += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 11
    for c in range(3, 3 + len(ALGOS)):
        ws.column_dimensions[get_column_letter(c)].width = 13
    return ws


def sheet_charts(wb):
    """One bar chart per (machine, algorithm), 8 total.

    X axis = the 8 precisions, ordered double/dd/td/qd | float/ds/ts/qs.
    bars   = compute time [s] @ n=1024 for each backend (linear Y axis).
    """
    ws = wb.create_sheet("Charts")
    ws["A1"] = ("Compute time [s] @ n=1024 (bars) - "
                "one chart per machine x algorithm")
    ws["A1"].font = TITLE_FONT
    # helper columns: A precision | serial t | simd1 t | simd2 t
    blk = 0
    for m in MACHINES:
        t = m["t"]
        for algo in ALGOS:
            base = 3 + blk * (len(PRECS) + 8)          # generous vertical spacing
            blk += 1
            ws.cell(row=base, column=1,
                    value=f"{m['key']} / {algo}").font = BOLD
            hdr_row = base + 1
            headers = ["precision"] + [f"{bk} [s]" for bk in m["backends"]]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=hdr_row, column=c, value=h)
            style_header(ws, hdr_row, 1, len(headers))
            for pi, p in enumerate(PRECS):
                rr = hdr_row + 1 + pi
                ws.cell(row=rr, column=1, value=p)
                ws.cell(row=rr, column=1).fill = MACH_FILL[m["key"]]
                for c, bk in enumerate(m["backends"], start=2):
                    tv = t.get((p, bk, algo, 1024))
                    cell = ws.cell(row=rr, column=c)
                    if tv is not None:
                        cell.value = round(tv, 9); cell.number_format = "0.000E+00"
            last = hdr_row + len(PRECS)

            nb = len(m["backends"])
            bar = BarChart(); bar.type = "col"; bar.grouping = "clustered"
            bar.title = (f"{m['key']}  {algo}  (n=1024): compute time [s]")
            # axis titles + make both axes explicit so Excel always renders the
            # labels and tick marks (openpyxl hides them unless delete=False)
            bar.x_axis.title = "precision"
            bar.y_axis.title = "compute time [s]"
            bar.x_axis.delete = False
            bar.y_axis.delete = False
            bar.x_axis.tickLblPos = "low"
            bdata = Reference(ws, min_col=2, max_col=1 + nb,
                              min_row=hdr_row, max_row=last)
            cats = Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=last)
            bar.add_data(bdata, titles_from_data=True); bar.set_categories(cats)
            bar.legend.position = "b"          # legend below the plot
            bar.height = 8.5; bar.width = 17
            ws.add_chart(bar, f"{get_column_letter(len(headers) + 2)}{base}")
    ws.column_dimensions["A"].width = 11
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 12
    return ws


def main():
    wb = Workbook()
    sheet_readme(wb)
    sheet_timing(wb)
    sheet_accuracy(wb)
    sheet_charts(wb)
    wb.save(XLSX)
    print("wrote", XLSX)
    for m in MACHINES:
        print(f"  {m['key']}: timing rows {len(m['t'])}, accuracy rows {len(m['a'])}")


if __name__ == "__main__":
    main()
