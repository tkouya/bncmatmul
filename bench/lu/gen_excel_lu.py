#!/usr/bin/env python3
"""Generate lu_fma_bench.xlsx from results/lu_fma_bench.csv.

Sheet "LU data"   : raw rows (prec, backend, fma, dim, lu_time, solve_time, maxrelerr)
Sheet "Speedup"   : per (prec, backend, dim): nofma vs fma LU time and speedup,
                    plus a bar chart of FMA speedups at the largest common dim.
"""
import csv, os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
CSV = os.path.join(HERE, "results", "lu_fma_bench.csv")
XLSX = os.path.join(os.path.dirname(os.path.dirname(HERE)), "lu_fma_bench.xlsx")

PRECS = ["dd", "td", "qd", "ds", "ts", "qs"]
BACKENDS = ["serial", "neon", "sve2"]

rows = []
with open(CSV) as f:
    for r in csv.DictReader(f):
        r["dim"] = int(r["dim"])
        r["lu_time"] = float(r["lu_time"])
        r["solve_time"] = float(r["solve_time"])
        r["maxrelerr"] = float(r["maxrelerr"])
        rows.append(r)

wb = Workbook()
hdr_font = Font(bold=True)
hdr_fill = PatternFill("solid", fgColor="DDE6F2")

ws = wb.active
ws.title = "LU data"
cols = ["prec", "backend", "fma", "dim", "lu_time", "solve_time", "maxrelerr"]
ws.append(cols)
for c in ws[1]:
    c.font = hdr_font; c.fill = hdr_fill
for r in rows:
    ws.append([r[k] for k in cols])
for w, col in zip((8, 9, 8, 7, 14, 14, 12), "ABCDEFG"):
    ws.column_dimensions[col].width = w

# ---- speedup sheet ----
ws2 = wb.create_sheet("Speedup")
ws2.append(["prec", "backend", "dim", "lu_nofma[s]", "lu_fma[s]", "lu_speedup",
            "solve_nofma[s]", "solve_fma[s]", "solve_speedup",
            "relerr_nofma", "relerr_fma"])
for c in ws2[1]:
    c.font = hdr_font; c.fill = hdr_fill

index = {(r["prec"], r["backend"], r["fma"], r["dim"]): r for r in rows}
dims = sorted({r["dim"] for r in rows})
for p in PRECS:
    for bk in BACKENDS:
        for d in dims:
            a = index.get((p, bk, "nofma", d))
            b = index.get((p, bk, "fma", d))
            if not a or not b:
                continue
            ws2.append([p, bk, d, a["lu_time"], b["lu_time"],
                        a["lu_time"] / b["lu_time"],
                        a["solve_time"], b["solve_time"],
                        a["solve_time"] / b["solve_time"],
                        a["maxrelerr"], b["maxrelerr"]])
for w, col in zip((8, 9, 7, 14, 14, 11, 14, 14, 13, 13, 13), "ABCDEFGHIJK"):
    ws2.column_dimensions[col].width = w

# ---- bar chart: speedup at dim=512 ----
CHART_DIM = 512
ws3 = wb.create_sheet("Chart512")
ws3.append(["prec/backend", "LU speedup", "solve speedup"])
for c in ws3[1]:
    c.font = hdr_font; c.fill = hdr_fill
for p in PRECS:
    for bk in BACKENDS:
        a = index.get((p, bk, "nofma", CHART_DIM))
        b = index.get((p, bk, "fma", CHART_DIM))
        if a and b:
            ws3.append([f"{p}-{bk}", a["lu_time"] / b["lu_time"],
                        a["solve_time"] / b["solve_time"]])
n = ws3.max_row
chart = BarChart()
chart.type = "col"
chart.title = f"LU / solve speedup by branch-free FMA+SIMD (dim={CHART_DIM})"
chart.y_axis.title = "speedup (nofma / fma)"
chart.x_axis.title = "precision - backend"
chart.height, chart.width = 10, 24
data = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=n)
cats = Reference(ws3, min_col=1, min_row=2, max_row=n)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws3.add_chart(chart, "D2")

# ---- ill-conditioned sheet (optional) ----
ILLCSV = os.path.join(HERE, "results", "lu_illcond.csv")
if os.path.exists(ILLCSV):
    ws4 = wb.create_sheet("IllCond")
    ws4.append(["family", "prec", "dim", "nofma(serial)", "fma(serial)",
                "fma(neon)", "fma(sve2)"])
    for c in ws4[1]:
        c.font = hdr_font; c.fill = hdr_fill
    irows = list(csv.DictReader(open(ILLCSV)))
    iidx = {(r["prec"], r["backend"], r["fma"], r["family"], int(r["dim"])):
            float(r["maxrelerr"]) for r in irows}
    for fam in ["hilbert", "frank"]:
        for p in PRECS:
            for d in range(2, 42, 2):
                key = (p, "serial", "nofma", fam, d)
                if key not in iidx:
                    continue
                ws4.append([fam, p, d, iidx[key],
                            iidx[(p, "serial", "fma", fam, d)],
                            iidx[(p, "neon", "fma", fam, d)],
                            iidx[(p, "sve2", "fma", fam, d)]])
    for w, col in zip((9, 7, 6, 14, 14, 14, 14), "ABCDEFG"):
        ws4.column_dimensions[col].width = w

wb.save(XLSX)
print("wrote", XLSX)
