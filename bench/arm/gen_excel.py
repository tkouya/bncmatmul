#!/usr/bin/env python3
"""Generate arm_bench.xlsx from dense_results.csv and spmv_results.csv.
Native (editable) Excel line charts: solid lines = time [s] (left, log axis),
dashed lines = speedup vs serial (right axis). A ManualLayout fixes the plot-area
box so the title (top), legend (bottom), x-axis values and all axis titles get
their own margins and never overlap."""
import csv, os
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "out")
XLSX = os.path.join(os.path.dirname(os.path.dirname(HERE)), "arm_bench.xlsx")

DIMS = [128, 256, 512, 1024, 2048]
REAL = ["double", "dd", "td", "qd", "float", "ds", "ts", "qs"]
CPLX = ["cd", "cf", "cdd", "ctd", "cqd", "cds", "cts", "cqs"]
OPS  = [("axpy", "axpy (add_mul: y = a*x + y)"),
        ("matvec", "matrix-vector product (y = A*x)"),
        ("matmul", "matrix product (C = A*B)")]
SPMV_REAL = ["d", "dd", "td", "qd"]
SPMV_FREAL = ["f", "ds", "ts", "qs"]
SPMV_CPLX = ["cd", "cdd", "ctd", "cqd"]
SPMV_FCPLX = ["cf", "cds", "cts", "cqs"]

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)


def load_dense():
    d = {}
    with open(os.path.join(OUT, "dense_results.csv")) as f:
        for r in csv.DictReader(f):
            d[(r["prec"], r["backend"], int(r["dim"]))] = {
                "axpy": float(r["axpy"]), "matvec": float(r["matvec"]),
                "matmul": float(r["matmul"])}
    return d


def load_spmv():
    d = {}
    with open(os.path.join(OUT, "spmv_results.csv")) as f:
        for r in csv.DictReader(f):
            d[(r["prec"], r["backend"])] = float(r["time"])
    return d


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def add_combo_chart(ws, title, hdr_row, last_row, ncat_col, t_first, t_last,
                    s_first, s_last, anchor, xtitle):
    """Editable line chart: time (log, left) + speedup (right), legend at bottom.
    ManualLayout reserves margins so labels/titles never overlap."""
    cats = Reference(ws, min_col=ncat_col, min_row=hdr_row + 1, max_row=last_row)

    tline = LineChart()
    tline.title = title
    tline.style = 2
    data = Reference(ws, min_col=t_first, max_col=t_last, min_row=hdr_row, max_row=last_row)
    tline.add_data(data, titles_from_data=True); tline.set_categories(cats)
    tline.y_axis.title = "time [s] (log)"
    tline.x_axis.title = xtitle
    tline.y_axis.scaling.logBase = 10
    tline.x_axis.delete = False
    tline.y_axis.delete = False
    tline.x_axis.tickLblPos = "low"
    tline.x_axis.majorTickMark = "out"
    tline.y_axis.majorTickMark = "out"
    for s in tline.series:
        s.smooth = False
        s.marker.symbol = "circle"; s.marker.size = 5
        s.graphicalProperties.line.width = 28000

    line = LineChart()
    ldata = Reference(ws, min_col=s_first, max_col=s_last, min_row=hdr_row, max_row=last_row)
    line.add_data(ldata, titles_from_data=True); line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.title = "speedup vs serial (x)"
    line.y_axis.crosses = "max"
    line.y_axis.delete = False
    line.y_axis.majorTickMark = "out"
    line.y_axis.majorGridlines = None   # no gridlines for the secondary (right) axis
    for s in line.series:
        s.smooth = False
        s.marker.symbol = "triangle"; s.marker.size = 6
        s.graphicalProperties.line.dashStyle = "dash"

    tline.y_axis.crosses = "autoZero"
    tline += line

    # explicit plot-area box: leave margins for title(top), y-titles(left/right),
    # x-values + x-title + legend(bottom)
    # plot box: small top (title) & bottom (x-values/x-title) margins so the
    # bottom legend sits close to the plot; side margins for the y-axis titles.
    tline.layout = Layout(manualLayout=ManualLayout(
        xMode="edge", yMode="edge", x=0.12, y=0.10, w=0.74, h=0.66))
    tline.legend.position = "b"
    tline.legend.overlay = False
    tline.width = 15.0
    tline.height = 10.0
    ws.add_chart(tline, anchor)


def write_table(ws, hdr_row, rows, first_col_head):
    heads = [first_col_head, "serial [s]", "neon [s]", "sve2 [s]",
             "neon/serial", "sve2/serial"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, len(heads))
    r = hdr_row + 1
    for key, ts, tn, tv in rows:
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=ts)
        ws.cell(row=r, column=3, value=tn)
        ws.cell(row=r, column=4, value=tv)
        ws.cell(row=r, column=5, value=(ts / tn) if (ts and tn) else None)
        ws.cell(row=r, column=6, value=(ts / tv) if (ts and tv) else None)
        for c in range(2, 5):
            ws.cell(row=r, column=c).number_format = "0.00E+00"
        for c in range(5, 7):
            ws.cell(row=r, column=c).number_format = "0.00"
        r += 1
    return r - 1


def dense_sheet(wb, dense, op, op_title, precs, group):
    ws = wb.create_sheet(f"{op}_{group}")
    ws["A1"] = f"{op_title}  —  {group} precisions"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("solid lines = time[s] (left, log);  dashed lines = speedup vs serial "
                "(right);  x-axis = N (axpy length = N*N).")
    ws["A2"].font = Font(italic=True, size=9)

    row = 4
    blocks = []
    for prec in precs:
        ws.cell(row=row, column=1, value=f"precision: {prec}").font = BOLD
        ws.cell(row=row, column=1).fill = SUB_FILL
        row += 1
        hdr_row = row
        trows = []
        for n in DIMS:
            ts = dense.get((prec, "serial", n), {}).get(op)
            tn = dense.get((prec, "neon", n), {}).get(op)
            tv = dense.get((prec, "sve2", n), {}).get(op)
            trows.append((n, ts, tn, tv))
        last = write_table(ws, hdr_row, trows, "N")
        blocks.append((prec, hdr_row, last))
        row = last + 2

    grid_cols = [8, 22]      # H and V (15cm chart ~9 cols + gap)
    grid_row0, grid_dr = 4, 24
    for i, (prec, hdr_row, last) in enumerate(blocks):
        anchor = f"{get_column_letter(grid_cols[i % 2])}{grid_row0 + (i // 2) * grid_dr}"
        add_combo_chart(ws, f"{prec}: {op}", hdr_row, last, 1, 2, 4, 5, 6, anchor, "N")
    for c, w in zip("ABCDEF", (6, 11, 11, 11, 11, 11)):
        ws.column_dimensions[c].width = w
    return ws


def spmv_sheet(wb, spmv, precs, group):
    ws = wb.create_sheet(f"spmv_{group}")
    ws["A1"] = f"sparse matrix-vector product (y = A*x)  —  {group} precisions"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("matrix: memplus.mtx (n=17758, nnz=126150).  solid = time[s] (left, log); "
                "dashed = speedup vs serial (right).")
    ws["A2"].font = Font(italic=True, size=9)
    hdr_row = 4
    trows = []
    for prec in precs:
        ts = spmv.get((prec, "serial")); tn = spmv.get((prec, "neon")); tv = spmv.get((prec, "sve2"))
        trows.append((prec, ts, tn, tv))
    last = write_table(ws, hdr_row, trows, "precision")
    add_combo_chart(ws, f"SpMV ({group})", hdr_row, last, 1, 2, 4, 5, 6, "H4", "precision")
    for c, w in zip("ABCDEF", (10, 12, 12, 12, 12, 12)):
        ws.column_dimensions[c].width = w


def raw_sheet(wb):
    for name, fn in (("raw_dense", "dense_results.csv"), ("raw_spmv", "spmv_results.csv")):
        ws = wb.create_sheet(name)
        p = os.path.join(OUT, fn)
        if not os.path.exists(p):
            continue
        with open(p) as f:
            for r, line in enumerate(csv.reader(f), 1):
                for c, v in enumerate(line, 1):
                    ws.cell(row=r, column=c, value=v)
        style_header(ws, 1, ws.max_column)


def main():
    dense = load_dense(); spmv = load_spmv()
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("README")
    notes = [
        ("ARM (NVIDIA DGX Spark / Grace, Neoverse V2) BNCmatmul benchmark", 14, True),
        ("", 10, False),
        ("Operations: axpy(add_mul), matrix-vector product, matrix product, sparse matrix-vector product", 10, False),
        ("Precisions (real): double, dd, td, qd, float, ds, ts, qs", 10, False),
        ("Precisions (complex): cd, cf, cdd, ctd, cqd, cds, cts, cqs", 10, False),
        ("  cf = native 'float _Complex' with HAND-WRITTEN SIMD kernels (bench/arm/cf_simd.h):", 10, False),
        ("       scalar / NEON (vld2q) / SVE2 (svld2) / AVX2 (addsub) / AVX-512 (fmaddsub).", 10, False),
        ("       NEON & SVE2 are built+validated+timed here; AVX2/AVX-512 are x86-only (compile elsewhere).", 10, False),
        ("Backends: serial (scalar lib), neon (Cortex-A76 NEON), sve2 (Neoverse-V2)", 10, False),
        ("", 10, False),
        ("Dense ops swept over N = 128, 256, 512, 1024, 2048 (axpy vector length = N*N).", 10, False),
        ("SpMV uses memplus.mtx (n=17758, nnz=126150).", 10, False),
        ("Real SpMV: double-based d/dd/td/qd and float-based f/ds/ts/qs (new; scalar+AVX2+AVX-512+NEON+SVE2).", 10, False),
        ("Complex SpMV: cd/cdd/ctd/cqd, float-based cds/cts/cqs ({re,im} pair of real DS/TS/QS sparse), and native cf.", 10, False),
        ("Float-based SpMV per-row padding uses the float SIMD width _BNC_S_WIDTH; SVE2 uses svcntw()-lane chunks.", 10, False),
        ("", 10, False),
        ("Each chart (native/editable): solid lines = time[s] on the left (log) axis;", 10, False),
        ("dashed lines = speedup vs serial (t_serial / t_backend) on the right axis;", 10, False),
        ("legend below the plot; x-axis shows N (or precision for SpMV).", 10, False),
    ]
    for i, (t, sz, b) in enumerate(notes, 1):
        ws.cell(row=i, column=1, value=t).font = Font(size=sz, bold=b)

    for op, title in OPS:
        dense_sheet(wb, dense, op, title, REAL, "real")
        dense_sheet(wb, dense, op, title, CPLX, "complex")
    spmv_sheet(wb, spmv, SPMV_REAL, "real")
    spmv_sheet(wb, spmv, SPMV_FREAL, "float")
    spmv_sheet(wb, spmv, SPMV_CPLX, "complex")
    spmv_sheet(wb, spmv, SPMV_FCPLX, "complex_float")
    raw_sheet(wb)

    wb.save(XLSX)
    print("wrote", XLSX, "| dense", len(dense), "spmv", len(spmv))


if __name__ == "__main__":
    main()
